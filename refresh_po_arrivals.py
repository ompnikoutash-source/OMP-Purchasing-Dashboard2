"""
PO Arrival Timeline — scheduled refresh script

Queries Gartman, writes results to an Excel file in the SharePoint
dashboard folder, and logs each run.  Run on a schedule via Windows
Task Scheduler using refresh_po_arrivals.bat.
"""

from __future__ import annotations

import os
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Allow imports from project root when run directly by Task Scheduler
sys.path.insert(0, str(Path(__file__).resolve().parent))
from core.db_connection import connect

# ── Configuration ──────────────────────────────────────────────────────────────
OUTPUT_DIR  = Path(r"C:\Users\niko\OneDrive - Old Master Products\Purchasing - Flooring Reports\Dashboard Files")
OUTPUT_FILE = "2026 Purchasing Report.xlsx"
SQL_FILE    = Path(__file__).resolve().parent / "po_arrival_timeline.sql"
LOG_FILE    = Path(__file__).resolve().parent / "refresh_po_arrivals.log"
MAX_LOG_LINES = 500   # rotate log after this many lines to keep it small

# ── Column widths (characters) ─────────────────────────────────────────────────
COL_WIDTHS = {
    "ETW":                     14,
    "Item Number":             16,
    "Description":             34,
    "Collection":              22,
    "PO Number":               12,
    "Physical Inv Available":  24,
    "Uncommitted on PO":       22,
    "Unattached Backorders":   24,
}

# ── Style constants ────────────────────────────────────────────────────────────
_THIN        = Side(style="thin")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_DATA_FONT   = Font(size=10, name="Calibri")
_ALT_FILL    = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
_TS_FONT     = Font(italic=True, size=9, color="888888", name="Calibri")

_NUM_COLS  = {"Physical Inv Available", "Uncommitted on PO", "Unattached Backorders"}
_DATE_COLS = {"ETW"}


# ── Logging ────────────────────────────────────────────────────────────────────

def _log(msg: str) -> None:
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    try:
        existing = LOG_FILE.read_text(encoding="utf-8").splitlines() if LOG_FILE.exists() else []
        if len(existing) >= MAX_LOG_LINES:
            existing = existing[-(MAX_LOG_LINES // 2):]   # keep newest half
        existing.append(line)
        LOG_FILE.write_text("\n".join(existing) + "\n", encoding="utf-8")
    except Exception:
        pass   # never let logging break the script


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    _log("=" * 60)
    _log("PO Arrival Timeline refresh started")

    # ── Read SQL ───────────────────────────────────────────────────────────────
    if not SQL_FILE.exists():
        _log(f"ERROR: SQL file not found: {SQL_FILE}")
        sys.exit(1)
    sql = SQL_FILE.read_text(encoding="utf-8")

    # ── Connect ────────────────────────────────────────────────────────────────
    try:
        conn = connect()
    except Exception as exc:
        _log(f"ERROR: DB connection failed: {exc}")
        sys.exit(1)

    # ── Query ──────────────────────────────────────────────────────────────────
    try:
        _log("Running query...")
        df = pd.read_sql(sql, conn)
        _log(f"Query complete — {len(df):,} rows returned")
    except Exception as exc:
        _log(f"ERROR: Query failed: {exc}")
        conn.close()
        sys.exit(1)
    finally:
        conn.close()

    # ── Clean up ETW column ────────────────────────────────────────────────────
    if "ETW" in df.columns:
        df["ETW"] = pd.to_datetime(df["ETW"], errors="coerce")
        df["ETW"] = df["ETW"].dt.date   # strip time portion; openpyxl writes as Excel date

    # ── Verify output directory ────────────────────────────────────────────────
    if not OUTPUT_DIR.exists():
        _log(f"ERROR: SharePoint sync folder not found: {OUTPUT_DIR}")
        _log("Make sure OneDrive is running and the folder is synced.")
        sys.exit(1)

    out_path = OUTPUT_DIR / OUTPUT_FILE

    # ── Write Excel (write to temp then replace to avoid OneDrive/Excel locks) ─
    tmp_path = out_path.with_suffix(".tmp.xlsx")
    try:
        _write_excel(df, tmp_path)
    except Exception as exc:
        _log(f"ERROR: Excel write failed: {exc}")
        sys.exit(1)

    # Replace target — retry a few times in case OneDrive briefly holds a lock
    for attempt in range(6):
        try:
            os.replace(tmp_path, out_path)
            _log(f"Saved: {out_path}")
            break
        except PermissionError as exc:
            if attempt < 5:
                _log(f"File locked (attempt {attempt+1}/6), retrying in 10s… ({exc})")
                time.sleep(10)
            else:
                _log(f"ERROR: Could not replace output file after 6 attempts: {exc}")
                _log("Make sure the file is not open in Excel.")
                sys.exit(1)

    _log("Refresh complete")


# ── Excel builder ──────────────────────────────────────────────────────────────

def _write_excel(df: pd.DataFrame, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Arrivals"

    columns   = list(df.columns)
    num_cols  = len(columns)
    num_rows  = len(df)

    # ── Header row ─────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(columns, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font       = _HEADER_FONT
        cell.fill       = _HEADER_FILL
        cell.border     = _BORDER
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_letter      = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col_name, 18)
    ws.row_dimensions[1].height = 28

    # ── Data rows ──────────────────────────────────────────────────────────────
    for row_offset, (_, row) in enumerate(df.iterrows()):
        r    = row_offset + 2
        fill = _ALT_FILL if row_offset % 2 == 0 else None

        for col_idx, col_name in enumerate(columns, start=1):
            raw   = row[col_name]
            value = None if (not isinstance(raw, str) and pd.isna(raw)) else raw

            cell        = ws.cell(row=r, column=col_idx, value=value)
            cell.font   = _DATA_FONT
            cell.border = _BORDER
            if fill:
                cell.fill = fill

            if col_name in _DATE_COLS:
                cell.number_format = "MM/DD/YYYY"
                cell.alignment     = Alignment(horizontal="center")
            elif col_name in _NUM_COLS:
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # ── Freeze header row, auto-filter ─────────────────────────────────────────
    ws.freeze_panes      = "A2"
    last_col             = get_column_letter(num_cols)
    ws.auto_filter.ref   = f"A1:{last_col}{num_rows + 1}"

    # ── Last-refreshed timestamp below data ────────────────────────────────────
    ts_row       = num_rows + 3
    ts_cell      = ws.cell(row=ts_row, column=1,
                           value=f"Last refreshed: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}")
    ts_cell.font = _TS_FONT

    wb.save(out_path)


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    main()
