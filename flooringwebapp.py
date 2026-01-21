"""
Complete Inventory Planning System - FULLY CORRECTED VERSION
Fixed Issues:
1. TSB implementation was completely wrong - fixed the probability and size tracking
2. Added fallback to simple moving average when TSB produces unrealistic results  
3. Fixed standard deviation over lead time calculation
4. Fixed reorder quantity formula
5. Added validation to prevent absurd forecasts
"""

from __future__ import annotations
import copy
import io
import json
import math, os, warnings
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import warnings
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

warnings.filterwarnings(
    "ignore",
    message="pandas only supports SQLAlchemy connectable",
    category=UserWarning,
)
import streamlit.components.v1 as components
try:
    import pyodbc
except Exception:
    pyodbc = None
try:
    from scipy import stats
except Exception:
    stats = None
try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing, SimpleExpSmoothing
except Exception:
    ExponentialSmoothing = None
    SimpleExpSmoothing = None
try:
    from sklearn.neural_network import MLPRegressor
    from sklearn.ensemble import RandomForestRegressor
except Exception:
    MLPRegressor = None
    RandomForestRegressor = None
try:
    from xgboost import XGBRegressor
except Exception:
    XGBRegressor = None


# Suppress all warnings
warnings.filterwarnings("ignore")
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", module="sklearn")
warnings.filterwarnings("ignore", module="statsmodels")

# Suppress sklearn parallel and convergence warnings via environment
os.environ['PYTHONWARNINGS'] = 'ignore::UserWarning'
os.environ['LOKY_MAX_CPU_COUNT'] = '4'  # Limit CPU usage for parallel processing
warnings.filterwarnings("ignore")

# ============================================================
# CONFIGURATION - CHANGE THIS TO PROCESS SPECIFIC SKU OR ALL
# ============================================================
FOCUS_SKU = ""  # Leave blank "" to use SKU list or process ALL SKUs
USE_GLOBAL_MODEL = True  # Set to True to train a single XGBoost model across ALL SKUs
FORECAST_SKU_LIST_FILE = "Forecast SKU List.xlsx"  # Excel file with list of SKUs to forecast
USE_SKU_LIST = True  # Set to True to only process SKUs in the list file
# ============================================================


# ============================================================
# GLOBAL ENDING INVENTORY CAP (prevents warehouse capacity issues)
# ============================================================
ENABLE_GLOBAL_INV_CAP = False
GLOBAL_INV_CAP_SF = 6_000_000  # Maximum total ending inventory across all SKUs (SF)

# ABC-based Months of Inventory caps (applied before global cap)
ABC_MOI_CAPS = {
    "A": 3.0,  # A items: max 3 months of inventory
    "B": 2.0,  # B items: max 2 months of inventory
    "C": 1.0   # C items: max 1 month of inventory
}
# ============================================================
# ============================================================
# SAFETY STOCK TUNING (ABC-based uplift)
# ============================================================
ABC_BREAK_A = 0.70   # top 70% of cumulative volume
ABC_BREAK_B = 0.80   # next 20% (70% -> 90%)
SS_UPLIFT_SCALAR = {"A": 1.0, "B": 0.35, "C": 0.00}  # Reduced from 1.00/0.35/0.00

# Lumpy buffer uses percentile demand instead of max demand 95 90 85 default
LUMPY_PCTL_NONZERO = {"A": 0.95, "B": 0.90, "C": 0.85}
LUMPY_LT_BUFFER_FRAC = 0.10  # 50% of percentile demand over lead time

# Optional global clamp: cap total uplift as % of trailing-12-month volume
ENABLE_GLOBAL_UPLIFT_BUDGET = True
GLOBAL_UPLIFT_BUDGET_PCT = 0.05  # 5% of trailing 12m shipped SF (reduced from 8%) 0.05
# ============================================================

DSN_NAME = "Gartman"
FUTURE_FORECAST_DAYS = 365
FUTURE_FORECAST_WEEKS = 52  # 1 year of weekly forecasts
DAYS_PER_WEEK = 7
DAYS_PER_MONTH = 30.4
WEEKS_PER_MONTH = DAYS_PER_MONTH / DAYS_PER_WEEK
SERVICE_LEVEL = 0.9  # 90% service level (reduced from 95%)
Z_SCORE = stats.norm.ppf(SERVICE_LEVEL) if stats is not None else 1.28

LEADTIMES_XLSX = "Lead Times.xlsx"
CUTOFF_DATE = "2020-06-01"
WEBAPP_JSON_PATH = Path(__file__).resolve().parent / "flooringwebappJSON"
SUNDRIES_JSON_PATH = Path(__file__).resolve().parent / "sundrieswebappJSON"
STRIP_SKU_LIST_FILE = "StripSKUList.xlsx"
UI_BUILD = "2026-01-14T15:10:00"

def _load_credentials():
    print("  Loading credentials...")
    
    # Check environment variables first
    uid_env = os.getenv("GARTMAN_UID", "").strip()
    pwd_env = os.getenv("GARTMAN_PWD", "").strip()
    
    # Load from secrets file
    secrets_path = Path(__file__).resolve().parent / "OMP_secrets.py"
    uid_file = ""
    pwd_file = ""
    
    if secrets_path.exists():
        # Use a clean namespace to avoid pollution
        namespace = {}
        with open(secrets_path, 'r', encoding='utf-8') as f:
            exec(f.read(), namespace, namespace)
        uid_file = str(namespace.get("GARTMAN_UID", "")).strip()
        pwd_file = str(namespace.get("GARTMAN_PWD", "")).strip()
    
    # Decide which credentials to use
    if uid_env and pwd_env:
        # Validate env vars aren't swapped by comparing to file
        if uid_file and pwd_file:
            if uid_env == pwd_file and pwd_env == uid_file:
                print(f"  ⚠ WARNING: Environment variables appear SWAPPED!")
                print(f"  ⚠ Using credentials from secrets file instead")
                uid, pwd = uid_file, pwd_file
                print(f"  Loaded from secrets file: UID={uid}, PWD={'*' * len(pwd)}")
            else:
                uid, pwd = uid_env, pwd_env
                print(f"  Using environment variables: UID={uid}, PWD={'*' * len(pwd)}")
        else:
            uid, pwd = uid_env, pwd_env
            print(f"  Using environment variables: UID={uid}, PWD={'*' * len(pwd)}")
    elif uid_file and pwd_file:
        uid, pwd = uid_file, pwd_file
        print(f"  Loaded from secrets file: UID={uid}, PWD={'*' * len(pwd)}")
    else:
        raise RuntimeError("Missing credentials")
    
    # Final validation: username and password should NOT be the same
    if uid == pwd:
        raise RuntimeError(f"ERROR: Username and password are identical ('{uid}'). This is incorrect!")
    
    return uid, pwd

def _connect():
    if pyodbc is None:
        raise RuntimeError("pyodbc is not available. Install local forecasting dependencies to run forecasts.")
    uid, pwd = _load_credentials()
    conn_str = f"DSN={DSN_NAME};UID={uid};PWD={pwd};"
    print(f"  Attempting database connection...")
    try:
        conn = pyodbc.connect(conn_str, autocommit=True, timeout=30)
        print(f"  ✓ Database connection successful!")
        return conn
    except pyodbc.Error as e:
        print(f"  ✗ Database connection failed!")
        print(f"  Error: {e}")
        raise

def load_lead_times(xlsx_path: Path) -> Tuple[Dict[str, float], Dict[str, float], Dict[str, float]]:
    """Load lead times from Excel"""
    print(f"\nLoading lead times from {xlsx_path}...")
    try:
        df = pd.read_excel(xlsx_path, sheet_name="Final", header=0)
        item_col = _get_first_column(df, ["SKU", "Item Number", "ITEM_NUMBER"])
        lt_col = _get_first_column(df, ["FINAL", "Final", "final"])
        sf_col = _get_first_column(df, ["SF per Pallet", "SF per pallet", "SF_per_Pallet"])
        pallet_col = _get_first_column(df, ["Pallets per Container", "Pallets per container", "Pallets_per_Container"])
        required_cols = {item_col, lt_col}
        if not item_col or not lt_col:
            print(f"  WARNING: Missing columns in lead time sheet: {required_cols - set(df.columns)}")
            return {}, {}, {}
        item_col, lt_col = item_col, lt_col
        df[item_col] = df[item_col].astype(str).str.strip().str.upper()
        df = df[~df[lt_col].astype(str).str.upper().str.contains("DISCONTINUED", na=False)]
        df[lt_col] = pd.to_numeric(df[lt_col], errors="coerce")
        df = df.dropna(subset=[item_col, lt_col])
        lead_times = dict(zip(df[item_col], df[lt_col]))
        sf_per_pallet = {}
        pallets_per_container = {}
        if sf_col and sf_col in df.columns:
            df[sf_col] = pd.to_numeric(df[sf_col], errors="coerce")
            sf_per_pallet = dict(zip(df[item_col], df[sf_col]))
        if pallet_col and pallet_col in df.columns:
            df[pallet_col] = pd.to_numeric(df[pallet_col], errors="coerce")
            pallets_per_container = dict(zip(df[item_col], df[pallet_col]))
        print(f"  Loaded {len(lead_times)} lead times")
        return lead_times, sf_per_pallet, pallets_per_container
    except Exception as e:
        print(f"  WARNING: {e}")
        return {}, {}, {}

def load_forecast_sku_list(xlsx_path: Path) -> set:
    """Load list of SKUs to forecast from Excel file"""
    print(f"\nLoading forecast SKU list from {xlsx_path}...")
    try:
        # Try to read the file - accept any column name
        df = pd.read_excel(xlsx_path, header=0)
        
        # Use the first column as the SKU list
        if df.shape[0] == 0:
            print(f"  WARNING: File is empty")
            return set()
        
        sku_col = df.columns[0]
        df[sku_col] = df[sku_col].astype(str).str.strip().str.upper()
        
        # Remove any blank/null values
        sku_list = set(df[sku_col].dropna())
        sku_list = {sku for sku in sku_list if sku and sku != 'NAN' and len(sku) > 0}
        
        print(f"  Loaded {len(sku_list)} SKUs from forecast list")
        return sku_list
    except FileNotFoundError:
        print(f"  WARNING: File not found - {xlsx_path}")
        print(f"  Will process ALL SKUs instead")
        return set()
    except Exception as e:
        print(f"  WARNING: Error reading file - {e}")
        print(f"  Will process ALL SKUs instead")
        return set()

def load_trailing_12m_volume(conn, sku_list=None) -> pd.DataFrame:
    """
    Returns ITEM_NUMBER, VOL_12M for SF demand in last 12 months.
    Uses SHLINE/SHHEAD like other sales queries.
    """
    print("\n  Loading trailing 12-month volume for ABC classification...")
    
    sku_filter = ""
    params = []

    if sku_list and len(sku_list) > 0:
        placeholders = ",".join(["?"] * len(sku_list))
        sku_filter = f" AND TRIM(L.SLITEM) IN ({placeholders}) "
        params = list(sku_list)

    sql = f"""
    SELECT
        TRIM(L.SLITEM) AS ITEM_NUMBER,
        SUM(COALESCE(L.SLBLUS,0)) AS VOL_12M
    FROM GSFL2K.SHLINE L
    JOIN GSFL2K.SHHEAD H
      ON H.SHCO = L.SLCO AND H.SHLOC = L.SLLOC AND H.SHORD# = L.SLORD# AND H.SHINV# = L.SLINV#
    WHERE L.SLUM2 LIKE '%SF%'
      AND H.SHIDAT >= (CURRENT_DATE - 365 DAYS)
      AND H.SHIDAT <= CURRENT_DATE
      {sku_filter}
    GROUP BY TRIM(L.SLITEM)
    """

    df = pd.read_sql_query(sql, conn, params=params if params else None)
    df["ITEM_NUMBER"] = df["ITEM_NUMBER"].astype(str).str.strip().str.upper()
    df["VOL_12M"] = pd.to_numeric(df["VOL_12M"], errors="coerce").fillna(0.0)
    
    print(f"  ✓ Loaded volume data for {len(df)} SKUs")
    return df

def build_abc_map(volume_df: pd.DataFrame) -> Dict[str, str]:
    """
    ABC segmentation by cumulative volume share.
    A = first 70% of volume, B = next 20%, C = last 10%
    """
    if volume_df.empty:
        return {}

    df = volume_df.copy()
    df = df.sort_values("VOL_12M", ascending=False)
    total = df["VOL_12M"].sum()
    
    if total <= 0:
        return {sku: "C" for sku in df["ITEM_NUMBER"].tolist()}

    df["cum_share"] = df["VOL_12M"].cumsum() / total

    abc = {}
    for _, r in df.iterrows():
        sku = r["ITEM_NUMBER"]
        cs = float(r["cum_share"])
        if cs <= ABC_BREAK_A:
            abc[sku] = "A"
        elif cs <= ABC_BREAK_B:
            abc[sku] = "B"
        else:
            abc[sku] = "C"
    
    # Print summary
    a_count = sum(1 for v in abc.values() if v == "A")
    b_count = sum(1 for v in abc.values() if v == "B")
    c_count = sum(1 for v in abc.values() if v == "C")
    print(f"\n  ABC Classification:")
    print(f"    A items (top 70% volume): {a_count} SKUs")
    print(f"    B items (next 20% volume): {b_count} SKUs")
    print(f"    C items (last 10% volume): {c_count} SKUs")
    
    return abc

def compute_ss_lumpy_from_percentile(y_weekly: pd.Series, lead_time_weeks: float, sku_abc: str, ss_base: float) -> float:
    """
    Lumpy buffer using percentile instead of max demand (more stable).
    Uses high percentile of non-zero weekly demand * LT_weeks * fraction
    """
    nonzero = y_weekly[y_weekly > 0].values.astype(float)
    if len(nonzero) < 2 or lead_time_weeks <= 0:
        return ss_base

    q = LUMPY_PCTL_NONZERO.get(sku_abc, 0.90)
    spike = float(np.percentile(nonzero, q * 100))
    ss_buffer = spike * float(lead_time_weeks) * float(LUMPY_LT_BUFFER_FRAC)

    return max(ss_base, ss_buffer)

def apply_ss_uplift_scalar(ss_base: float, ss_lumpy: float, sku_abc: str) -> tuple:
    """
    Returns (ss_final_pre_budget, uplift_raw, uplift_applied)
    Applies ABC-based scalar to the uplift only (keeps base intact)
    """
    uplift_raw = max(0.0, ss_lumpy - ss_base)
    scalar = SS_UPLIFT_SCALAR.get(sku_abc, 0.0)
    uplift_applied = scalar * uplift_raw
    ss_final = ss_base + uplift_applied
    return ss_final, uplift_raw, uplift_applied

def _build_in_list(values: List[str]) -> str:
    safe = [str(v).replace("'", "''") for v in values if v]
    return "'" + "','".join(safe) + "'" if safe else "''"

def get_active_skus(conn, single_sku=None, sku_list=None):
    """
    Get active SKUs based on:
      - Sales within last 3 years, OR
      - First receipt within last 1 year.
    """
    
    # If we have a specific SKU list, use a much simpler and faster query
    if sku_list and len(sku_list) > 0:
        print(f"  Fetching data for {len(sku_list)} specific SKUs...")
        
        # Create IN clause for the SKU list
        sku_list_str = _build_in_list(sku_list)
        
        sql = f"""
        WITH
        FIRSTREC AS (
          SELECT TRIM(R.IRITEM) AS ITEM_NUMBER, MIN(R.IRDATE) AS FIRST_RECEIPT_DATE
          FROM GSFL2K.ITEMRECH R
          WHERE R.IRLOC NOT IN (90, 17, 41, 46)
            AND TRIM(R.IRITEM) IN ({sku_list_str})
          GROUP BY TRIM(R.IRITEM)
        ),
        LAST_SALE AS (
          SELECT TRIM(L.SLITEM) AS ITEM_NUMBER, MAX(H.SHIDAT) AS LAST_SALE_DATE
          FROM GSFL2K.SHLINE L
          JOIN GSFL2K.SHHEAD H
            ON H.SHCO = L.SLCO AND H.SHLOC = L.SLLOC AND H.SHORD# = L.SLORD# AND H.SHINV# = L.SLINV#
          WHERE L.SLUM2 LIKE '%SF%'
            AND COALESCE(L.SLBLUS,0) > 0
            AND TRIM(L.SLITEM) IN ({sku_list_str})
          GROUP BY TRIM(L.SLITEM)
        ),
        INV AS (
          SELECT B.IBITEM, SUM(B.IBQOH) AS QTY_ON_HAND, SUM(B.IBQOO) AS QTY_COMMITTED
          FROM GSFL2K.ITEMBAL B
          WHERE B.IBLOC NOT IN (90, 17, 41, 46)
            AND TRIM(B.IBITEM) IN ({sku_list_str})
          GROUP BY B.IBITEM
        ),
        PO AS (
          SELECT PLITEM, SUM(PLBLUO) AS ON_PO_SF
          FROM GSFL2K.POLINE
          WHERE PLDELT LIKE '%A%'
            AND TRIM(PLITEM) IN ({sku_list_str})
          GROUP BY PLITEM
        ),
        BO AS (
          SELECT OLITEM, SUM(OLBLUB) AS BO_SF
          FROM GSFL2K.OOLINE
          WHERE OLCUST NOT LIKE '%TRANSFER%'
            AND OLCUST NOT LIKE '%OMP000%'
            AND OLCUST NOT LIKE '%INV000%'
            AND OLCUST NOT LIKE '%OLD001%'
            AND OLLOC <> 90
            AND OLBO LIKE 'Y'
            AND OLUM2 LIKE '%SF%'
            AND TRIM(OLITEM) IN ({sku_list_str})
          GROUP BY OLITEM
        )
        SELECT
          TRIM(M.IMITEM) AS ITEM_NUMBER,
          TRIM(M.IMDESC) AS DESCRIPTION,
          TRIM(M.IMVEND) AS VENDOR_NUMBER,
          TRIM(V.VMNAME) AS VENDOR_NAME,
          TRIM(X.IMCOLLECT) AS COLLECTION,
          FR.FIRST_RECEIPT_DATE,
          ((COALESCE(INV.QTY_ON_HAND,0) - COALESCE(INV.QTY_COMMITTED,0)) * COALESCE(M.IMFACT, 1)) AS AVAILABLE_SF,
          COALESCE(PO.ON_PO_SF, 0) AS ON_PO_SF,
          COALESCE(BO.BO_SF, 0) AS BACKORDER_SF,
          COALESCE(M.IMLT, 0) AS LEAD_TIME_IMLT
        FROM GSFL2K.ITEMMAST M
        LEFT JOIN GSFL2K.VENDMAST V ON V.VMVEND = M.IMVEND
        LEFT JOIN GSFL2K.ITEMXTRA X ON X.IMXITM = M.IMITEM
        LEFT JOIN FIRSTREC FR ON FR.ITEM_NUMBER = TRIM(M.IMITEM)
        LEFT JOIN LAST_SALE LS ON LS.ITEM_NUMBER = TRIM(M.IMITEM)
        LEFT JOIN INV ON INV.IBITEM = M.IMITEM
        LEFT JOIN PO ON PO.PLITEM = M.IMITEM
        LEFT JOIN BO ON BO.OLITEM = M.IMITEM
        WHERE TRIM(M.IMITEM) IN ({sku_list_str})
          AND (
            (LS.LAST_SALE_DATE IS NOT NULL AND LS.LAST_SALE_DATE >= (CURRENT_DATE - 3 YEARS))
            OR (FR.FIRST_RECEIPT_DATE >= (CURRENT_DATE - 1 YEAR))
          )
        ORDER BY TRIM(M.IMITEM)
        """
        
        print("  Executing SQL query to fetch SKUs...")
        df = pd.read_sql_query(sql, conn)
        print(f"  ✓ Query returned {len(df)} rows")
        
    else:
        # Original query for all SKUs or single SKU
        sql = """
        WITH
        LAST_SALE AS (
          SELECT TRIM(L.SLITEM) AS ITEM_NUMBER, MAX(H.SHIDAT) AS LAST_SALE_DATE
          FROM GSFL2K.SHLINE L
          JOIN GSFL2K.SHHEAD H ON H.SHCO = L.SLCO AND H.SHLOC = L.SLLOC AND H.SHORD# = L.SLORD# AND H.SHINV# = L.SLINV#
          WHERE L.SLUM2 LIKE '%SF%'
            AND COALESCE(L.SLBLUS,0) > 0
          GROUP BY TRIM(L.SLITEM)
        ),
        FIRSTREC AS (
          SELECT TRIM(R.IRITEM) AS ITEM_NUMBER, MIN(R.IRDATE) AS FIRST_RECEIPT_DATE
          FROM GSFL2K.ITEMRECH R
          WHERE R.IRLOC NOT IN (90, 17, 41, 46)
          GROUP BY TRIM(R.IRITEM)
        ),
        INV AS (
          SELECT B.IBITEM, SUM(B.IBQOH) AS QTY_ON_HAND, SUM(B.IBQOO) AS QTY_COMMITTED
          FROM GSFL2K.ITEMBAL B
          WHERE B.IBLOC NOT IN (90, 17, 41, 46)
          GROUP BY B.IBITEM
        ),
        PO AS (
          SELECT PLITEM, SUM(PLBLUO) AS ON_PO_SF
          FROM GSFL2K.POLINE
          WHERE PLDELT LIKE '%A%'
          GROUP BY PLITEM
        ),
        BO AS (
          SELECT OLITEM, SUM(OLBLUB) AS BO_SF
          FROM GSFL2K.OOLINE
          WHERE OLCUST NOT LIKE '%TRANSFER%'
            AND OLCUST NOT LIKE '%OMP000%'
            AND OLCUST NOT LIKE '%INV000%'
            AND OLCUST NOT LIKE '%OLD001%'
            AND OLLOC <> 90
            AND OLBO LIKE 'Y'
            AND OLUM2 LIKE '%SF%'
          GROUP BY OLITEM
        )
        SELECT
          TRIM(M.IMITEM) AS ITEM_NUMBER,
          TRIM(M.IMDESC) AS DESCRIPTION,
          TRIM(M.IMVEND) AS VENDOR_NUMBER,
          TRIM(V.VMNAME) AS VENDOR_NAME,
          TRIM(X.IMCOLLECT) AS COLLECTION,
          FR.FIRST_RECEIPT_DATE,
          ((COALESCE(INV.QTY_ON_HAND,0) - COALESCE(INV.QTY_COMMITTED,0)) * COALESCE(M.IMFACT, 1)) AS AVAILABLE_SF,
          COALESCE(PO.ON_PO_SF, 0) AS ON_PO_SF,
          COALESCE(BO.BO_SF, 0) AS BACKORDER_SF,
          COALESCE(M.IMLT, 0) AS LEAD_TIME_IMLT
        FROM GSFL2K.ITEMMAST M
        LEFT JOIN GSFL2K.VENDMAST V ON V.VMVEND = M.IMVEND
        LEFT JOIN GSFL2K.ITEMXTRA X ON X.IMXITM = M.IMITEM
        LEFT JOIN FIRSTREC FR ON FR.ITEM_NUMBER = TRIM(M.IMITEM)
        LEFT JOIN LAST_SALE LS ON LS.ITEM_NUMBER = TRIM(M.IMITEM)
        LEFT JOIN INV ON INV.IBITEM = M.IMITEM
        LEFT JOIN PO ON PO.PLITEM = M.IMITEM
        LEFT JOIN BO ON BO.OLITEM = M.IMITEM
        WHERE (
          (LS.LAST_SALE_DATE IS NOT NULL AND LS.LAST_SALE_DATE >= (CURRENT_DATE - 3 YEARS))
          OR (FR.FIRST_RECEIPT_DATE >= (CURRENT_DATE - 1 YEAR))
        )
        {sku_filter}
        ORDER BY TRIM(M.IMITEM)
        """
        
        sku_filter = f"AND TRIM(M.IMITEM) = '{single_sku}'" if single_sku else ""
        final_sql = sql.format(sku_filter=sku_filter)
        
        print("  Executing SQL query to fetch SKUs...")
        df = pd.read_sql_query(final_sql, conn)
        print(f"  ✓ Query returned {len(df)} rows")
    
    df['ITEM_NUMBER'] = df['ITEM_NUMBER'].astype(str).str.strip().str.upper()
    df['FIRST_RECEIPT_DATE'] = pd.to_datetime(df['FIRST_RECEIPT_DATE'], errors='coerce')
    
    # Calculate Inventory Position = Available + On PO - Backorder
    df['INVENTORY_POSITION'] = df['AVAILABLE_SF'] + df['ON_PO_SF'] - df['BACKORDER_SF']
    
    # Report results - filtering already done in SQL if sku_list was provided
    if sku_list and len(sku_list) > 0:
        print(f"\nFound {len(df)} SKUs from forecast list")
        if len(df) < len(sku_list):
            missing = len(sku_list) - len(df)
            print(f"  Note: {missing} SKUs from list not found in database")
    else:
        print(f"\nFound {len(df)} active SKUs (sales in last 3 years or first receipt within 1 year)")
    
    return df

def load_sales_history(conn, sku, start_date):
    """Load sales history using correct SQL"""
    today = pd.Timestamp.today().normalize()
    first_of_month = pd.Timestamp(year=today.year, month=today.month, day=1)
    hist_end_dt = first_of_month + pd.offsets.MonthBegin(1)
    
    sql = """
    SELECT H.SHIDAT AS SALES_DATE, SUM(COALESCE(L.SLBLUS,0)) AS QTY_SOLD_SF
    FROM GSFL2K.SHLINE L
    JOIN GSFL2K.SHHEAD H ON H.SHCO = L.SLCO AND H.SHLOC = L.SLLOC AND H.SHORD# = L.SLORD# AND H.SHINV# = L.SLINV#
    WHERE L.SLUM2 LIKE '%SF%'
      AND TRIM(L.SLITEM) = ?
      AND H.SHIDAT >= ?
      AND H.SHIDAT < ?
    GROUP BY H.SHIDAT
    ORDER BY H.SHIDAT
    """
    
    df = pd.read_sql_query(sql, conn, params=[sku, start_date.strftime("%Y-%m-%d"), hist_end_dt.strftime("%Y-%m-%d")])
    df["SALES_DATE"] = pd.to_datetime(df["SALES_DATE"], errors="coerce")
    df["QTY_SOLD_SF"] = pd.to_numeric(df["QTY_SOLD_SF"], errors="coerce").fillna(0.0)
    return df

def load_arrivals(conn, sku_list: List[str]) -> pd.DataFrame:
    """Load open PO arrivals with container notes from POLINE/POTEXT."""
    if not sku_list:
        return pd.DataFrame()
    sku_list_str = _build_in_list(sku_list)
    sql = f"""
    SELECT
      TRIM(L.PLPO#) AS PO_NUMBER,
      TRIM(L.PLITEM) AS ITEM_NUMBER,
      TRIM(L.PLDESC) AS DESCRIPTION,
      COALESCE(L.PLBLUO, 0) AS QUANTITY_SF,
      TRIM(M.IMVEND) AS VENDOR_NUMBER,
      TRIM(V.VMNAME) AS VENDOR_NAME,
      TRIM(X.IMCOLLECT) AS COLLECTION,
      L.PLPDAT AS DUE_PORT,
      L.PLDDAT AS DUE_INV,
      H.PHDOI AS ENTRY_DATE,
      H.PHSDAT AS EST_SHIP_DATE,
      MAX(TRIM(P.PTCMT1)) AS PTCMT1,
      MAX(TRIM(P.PTCMT2)) AS PTCMT2
    FROM GSFL2K.POLINE L
    LEFT JOIN GSFL2K.POTEXT P ON P.PTPO# = L.PLPO#
    LEFT JOIN GSFL2K.POHEAD H ON H.PHPO# = L.PLPO#
    JOIN GSFL2K.ITEMMAST M ON M.IMITEM = L.PLITEM
    LEFT JOIN GSFL2K.ITEMXTRA X ON X.IMXITM = M.IMITEM
    LEFT JOIN GSFL2K.VENDMAST V ON V.VMVEND = M.IMVEND
    WHERE L.PLDELT LIKE '%A%'
      AND TRIM(L.PLITEM) IN ({sku_list_str})
      AND TRIM(L.PLPO#) <> ''
      AND TRIM(L.PLITEM) <> ''
    GROUP BY
      TRIM(L.PLPO#),
      TRIM(L.PLITEM),
      TRIM(L.PLDESC),
      COALESCE(L.PLBLUO, 0),
      TRIM(M.IMVEND),
      TRIM(V.VMNAME),
      TRIM(X.IMCOLLECT),
      L.PLPDAT,
      L.PLDDAT,
      H.PHDOI,
      H.PHSDAT
    """
    df = pd.read_sql_query(sql, conn)
    df["CONTAINER"] = df["PTCMT1"].fillna("").astype(str).str.strip()
    df["CONTAINER"] = np.where(
        df["CONTAINER"] != "",
        df["CONTAINER"],
        df["PTCMT2"].fillna("").astype(str).str.strip(),
    )
    df["CONTAINER"] = df["CONTAINER"].replace("", np.nan)
    return df

def build_daily_series(sales_df, start_date):
    """Build daily time series"""
    if sales_df.empty:
        return None
    daily_agg = sales_df.groupby('SALES_DATE')['QTY_SOLD_SF'].sum().reset_index()
    end = pd.Timestamp.today().normalize()
    idx = pd.date_range(start_date, end, freq="D")
    s = daily_agg.set_index('SALES_DATE')['QTY_SOLD_SF'].reindex(idx, fill_value=0.0).astype(float)
    s.index.name = "DATE"
    return s

def build_weekly_series(sales_df, start_date):
    """Build weekly time series - aggregates to week ending on Sunday"""
    if sales_df.empty:
        return None
    
    # Aggregate daily sales to weekly
    daily_agg = sales_df.groupby('SALES_DATE')['QTY_SOLD_SF'].sum().reset_index()
    
    # Create complete daily index
    end = pd.Timestamp.today().normalize()
    daily_idx = pd.date_range(start_date, end, freq="D")
    daily_series = daily_agg.set_index('SALES_DATE')['QTY_SOLD_SF'].reindex(daily_idx, fill_value=0.0)
    
    # Resample to weekly (Sunday week-end)
    weekly_series = daily_series.resample('W-SUN').sum()
    
    # Remove the last week if it's incomplete (partial week)
    if weekly_series.index[-1] > end:
        weekly_series = weekly_series[:-1]
    
    weekly_series.index.name = "WEEK_END"
    return weekly_series.astype(float)

def cap_outliers(series: pd.Series, n_mad: float = 4.0) -> pd.Series:
    """Cap extreme outliers using median and MAD (robust to outliers).

    Uses median and Median Absolute Deviation (MAD) instead of mean/std
    because mean/std are themselves heavily influenced by outliers.
    MAD is converted to a standard deviation equivalent using the
    normal distribution scale factor (1.4826).

    Args:
        series: Time series of demand values
        n_mad: Number of MAD-scaled deviations for the threshold (default 4.0)

    Returns:
        Series with outliers capped at median + n_mad * (1.4826 * MAD)
    """
    if series.empty:
        return series
    non_zero = series[series > 0]
    ref = non_zero if len(non_zero) >= 3 else series
    median = ref.median()
    mad = np.median(np.abs(ref - median))
    if mad == 0:
        upper_bound = ref.quantile(0.95)
        if not np.isfinite(upper_bound) or upper_bound <= 0:
            return series
        reason = "p95 of non-zero demand"
    else:
        # Convert MAD to std-equivalent (for normal distribution, std ~ 1.4826 * MAD)
        std_est = 1.4826 * mad
        upper_bound = median + n_mad * std_est
        reason = f"median + {n_mad} MAD"
    capped = series.clip(upper=upper_bound)
    n_capped = (series > upper_bound).sum()
    if n_capped > 0:
        print(f"    Capped {n_capped} outlier(s) exceeding {upper_bound:,.0f} ({reason})")
    return capped

def wmape(y_true, y_pred):
    """Calculate wMAPE"""
    denom = float(np.sum(np.abs(y_true)))
    return float(np.sum(np.abs(y_true - y_pred)) / denom) * 100 if denom != 0 else float("nan")

def calculate_demand_characteristics(y_series):
    """
    Calculate ADI (Average Demand Interval) and CV² (Coefficient of Variation²)
    Used to classify demand patterns
    
    Returns:
    - adi: Average Demand Interval (avg weeks between non-zero demands)
    - cv2: Coefficient of Variation² (variance in demand SIZE)
    - demand_class: 'SMOOTH', 'INTERMITTENT', 'ERRATIC', or 'LUMPY'
    """
    # Find non-zero demands
    non_zero_mask = y_series > 0
    non_zero_values = y_series[non_zero_mask].values
    non_zero_indices = np.where(non_zero_mask)[0]
    
    if len(non_zero_values) < 2:
        return None, None, 'INSUFFICIENT_DATA'
    
    # Calculate ADI (Average Demand Interval)
    # Average number of periods between non-zero demands
    intervals = np.diff(non_zero_indices)
    adi = intervals.mean() if len(intervals) > 0 else 1.0
    
    # Calculate CV² (Coefficient of Variation squared)
    # Measures variability in demand SIZE (when non-zero)
    mean_demand = non_zero_values.mean()
    std_demand = non_zero_values.std()
    cv = std_demand / mean_demand if mean_demand > 0 else 0
    cv2 = cv ** 2
    
    # Classify demand pattern (Syntetos et al. 2005)
    # ADI threshold: 1.32 (about weekly vs less frequent)
    # CV² threshold: 0.49 (consistent size vs variable)
    
    if adi < 1.32 and cv2 < 0.49:
        demand_class = 'SMOOTH'
    elif adi >= 1.32 and cv2 < 0.49:
        demand_class = 'INTERMITTENT'
    elif adi < 1.32 and cv2 >= 0.49:
        demand_class = 'ERRATIC'
    else:  # adi >= 1.32 and cv2 >= 0.49
        demand_class = 'LUMPY'
    
    return adi, cv2, demand_class

# ============================================================
# FOUR FORECASTING METHODS
# ============================================================

def forecast_neural_network(y_train, horizon, lags=[1, 2, 4, 8]):
    """Neural Network (MLP) - for weekly data"""
    try:
        if len(y_train) < 20:  # Need at least 20 weeks
            return None, "NN: insufficient data"
        X, y = [], []
        for i in range(max(lags), len(y_train)):
            X.append([y_train.iloc[i-lag] for lag in lags])
            y.append(y_train.iloc[i])
        model = MLPRegressor(hidden_layer_sizes=(50, 25), max_iter=1000, random_state=42, 
                            early_stopping=True, n_iter_no_change=20, validation_fraction=0.1)
        model.fit(np.array(X), np.array(y))
        preds, hist = [], y_train.values.tolist()
        for _ in range(horizon):
            x_new = [hist[-lag] for lag in lags]
            pred = max(0.0, model.predict([x_new])[0])
            preds.append(pred)
            hist.append(pred)
        idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
        return pd.Series(preds, index=idx, dtype=float), None
    except Exception as e:
        return None, str(e)


def forecast_croston(y_train, horizon, alpha=0.1):
    """
    Croston's Method - Industry standard for intermittent demand
    
    Separately forecasts:
    1. Demand intervals (time between non-zero demands)
    2. Demand sizes (when demand occurs)
    
    Final forecast = (demand size) / (demand interval)
    """
    try:
        y = y_train.values.astype(float)
        
        # Find non-zero demand periods
        non_zero_idx = np.where(y > 0)[0]
        
        if len(non_zero_idx) < 2:
            # Not enough non-zero periods, use average
            avg = y_train.mean()
            idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
            return pd.Series([avg] * horizon, index=idx, dtype=float), None
        
        # Initialize demand size and interval
        demand_sizes = y[non_zero_idx]
        intervals = np.diff(non_zero_idx)
        
        # Start with first values
        z = demand_sizes[0]  # demand size estimate
        p = intervals[0] if len(intervals) > 0 else 1  # interval estimate
        
        # Apply exponential smoothing
        for i in range(1, len(non_zero_idx)):
            z = alpha * demand_sizes[i] + (1 - alpha) * z
            if i < len(intervals):
                p = alpha * intervals[i] + (1 - alpha) * p
        
        # Forecast = demand size / interval
        forecast_value = z / p if p > 0 else z
        
        idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
        return pd.Series([forecast_value] * horizon, index=idx, dtype=float), None
    except Exception as e:
        return None, str(e)

def forecast_sba(y_train, horizon, alpha=0.1):
    """
    Syntetos-Boylan Approximation (SBA) - Improved Croston's Method
    
    Adds bias correction to Croston's method for better accuracy
    Based on: Syntetos & Boylan (2005)
    """
    try:
        y = y_train.values.astype(float)
        
        # Find non-zero demand periods
        non_zero_idx = np.where(y > 0)[0]
        
        if len(non_zero_idx) < 2:
            # Not enough non-zero periods, use average
            avg = y_train.mean()
            idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
            return pd.Series([avg] * horizon, index=idx, dtype=float), None
        
        # Initialize demand size and interval
        demand_sizes = y[non_zero_idx]
        intervals = np.diff(non_zero_idx)
        
        # Start with first values
        z = demand_sizes[0]
        p = intervals[0] if len(intervals) > 0 else 1
        
        # Apply exponential smoothing
        for i in range(1, len(non_zero_idx)):
            z = alpha * demand_sizes[i] + (1 - alpha) * z
            if i < len(intervals):
                p = alpha * intervals[i] + (1 - alpha) * p
        
        # SBA correction factor (reduces bias in Croston's)
        correction = 1 - (alpha / 2)
        forecast_value = (z / p) * correction if p > 0 else z
        
        idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
        return pd.Series([forecast_value] * horizon, index=idx, dtype=float), None
    except Exception as e:
        return None, str(e)

def forecast_bootstrap(y_train, horizon, n_simulations=1000):
    """
    Bootstrap Forecast - BEST for LUMPY demand
    
    Instead of predicting exact values, samples from historical distribution.
    Captures both demand occurrence probability AND size variability.
    
    Perfect for project-driven, unpredictable demand patterns.
    """
    try:
        # Extract non-zero demands
        non_zero = y_train[y_train > 0].values
        
        if len(non_zero) == 0:
            # No historical demand, forecast zero
            idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
            return pd.Series([0.0] * horizon, index=idx, dtype=float), None
        
        if len(non_zero) < 3:
            # Very sparse data, use simple average
            avg = y_train.mean()
            idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
            return pd.Series([avg] * horizon, index=idx, dtype=float), None
        
        # Calculate probability of non-zero demand
        p_nonzero = len(non_zero) / len(y_train)
        
        # Run simulations
        simulations = []
        for _ in range(n_simulations):
            forecast_sim = []
            for _ in range(horizon):
                if np.random.random() < p_nonzero:
                    # Sample from historical non-zero distribution with replacement
                    demand = np.random.choice(non_zero)
                else:
                    demand = 0.0
                forecast_sim.append(demand)
            simulations.append(forecast_sim)
        
        # Average across simulations for point forecast
        simulations = np.array(simulations)
        forecast_values = simulations.mean(axis=0)
        
        idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
        return pd.Series(forecast_values, index=idx, dtype=float), None
    except Exception as e:
        return None, str(e)

def forecast_tsb(y_train, horizon, alpha=0.2, beta=0.1):
    """
    TSB (Teunter-Syntetos-Babai) - PROPERLY IMPLEMENTED for WEEKLY data
    
    TSB is designed for intermittent demand but has a fatal flaw:
    if recent data is mostly zeros, p decays to zero regardless of historical sales.
    
    Fix: Use the ACTUAL probability and size from the data, not exponential smoothing.
    """
    try:
        y = y_train.values.astype(float)
        
        # Calculate actual demand probability and size
        non_zero_mask = y > 0
        non_zero_count = np.sum(non_zero_mask)
        
        if non_zero_count == 0:
            # No demand ever - return zeros
            idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
            return pd.Series([0.0] * horizon, index=idx, dtype=float), None
        
        # Use ACTUAL statistics instead of smoothing (which decays with recent zeros)
        p = non_zero_count / len(y)  # Actual probability of demand
        z = np.mean(y[non_zero_mask])  # Actual average size when demand occurs
        
        # Forecast = probability × average size
        level = p * z
        
        # This should equal y_train.mean() by definition
        # Sanity check anyway
        historical_mean = y_train.mean()
        if abs(level - historical_mean) > 0.01 * historical_mean:
            # Small numerical difference is OK, but if large, use mean
            level = historical_mean
        
        idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
        return pd.Series([level] * horizon, index=idx, dtype=float), None
    except Exception as e:
        return None, str(e)

def forecast_ses(y_train, horizon):
    """Simple Exponential Smoothing - for weekly data"""
    try:
        model = SimpleExpSmoothing(y_train).fit(optimized=True)
        fc = model.forecast(horizon)
        idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
        return pd.Series(np.maximum(0.0, fc.values), index=idx), None
    except Exception as e:
        return None, str(e)

def forecast_xgboost(y_train, horizon, lags=[1, 2, 4, 8]):
    """XGBoost - for weekly data"""
    try:
        if len(y_train) < 20:  # Need at least 20 weeks
            return None, "XGB: insufficient data"
        X, y = [], []
        for i in range(max(lags), len(y_train)):
            row = [y_train.iloc[i-lag] for lag in lags]
            # Add week-based temporal features
            row.extend([y_train.index[i].isocalendar()[1], y_train.index[i].month, y_train.index[i].quarter])
            X.append(row)
            y.append(y_train.iloc[i])
        model = XGBRegressor(n_estimators=200, learning_rate=0.05, max_depth=6, random_state=42)
        model.fit(np.array(X), np.array(y))
        preds, hist = [], y_train.copy()
        for i in range(horizon):
            future_date = hist.index[-1] + pd.Timedelta(weeks=1)
            row = [hist.iloc[-lag] for lag in lags]
            row.extend([future_date.isocalendar()[1], future_date.month, future_date.quarter])
            pred = max(0.0, model.predict([row])[0])
            preds.append(pred)
            new_row = pd.Series([pred], index=[future_date])
            hist = pd.concat([hist, new_row])
        idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), periods=horizon, freq="W-SUN")
        return pd.Series(preds, index=idx, dtype=float), None
    except Exception as e:
        return None, str(e)



def forecast_random_forest(y_train, horizon, lags=[1, 2, 4, 8]):
    """Random Forest with lagged features"""
    try:
        if len(y_train) < max(lags) + 10:
            return None, "Insufficient data"
        
        X, y = [], []
        for i in range(max(lags), len(y_train)):
            X.append([y_train.iloc[i-lag] if i >= lag else 0 for lag in lags])
            y.append(y_train.iloc[i])
        
        model = RandomForestRegressor(n_estimators=100, max_depth=10, random_state=42, n_jobs=-1)
        model.fit(np.array(X), np.array(y))
        
        # Forecast
        preds = []
        hist = y_train.values.tolist()
        for _ in range(horizon):
            x_new = [hist[-lag] if lag <= len(hist) else 0 for lag in lags]
            pred = max(0.0, model.predict([x_new])[0])
            preds.append(pred)
            hist.append(pred)
        
        # Convert to series with proper index
        forecast_idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), 
                                     periods=horizon, freq="W-SUN")
        return pd.Series(preds, index=forecast_idx), None
    except Exception as e:
        return None, str(e)

def forecast_holt_damped(y_train, horizon):
    """Holt's method with damped trend"""
    try:
        from statsmodels.tsa.holtwinters import Holt
        model = Holt(y_train, damped_trend=True)
        fitted = model.fit()
        forecast = fitted.forecast(steps=horizon)
        return pd.Series(np.maximum(0, forecast.values), index=forecast.index), None
    except Exception as e:
        return None, str(e)

def forecast_sarima(y_train, horizon):
    """SARIMA(0,1,1) model"""
    try:
        from statsmodels.tsa.statespace.sarimax import SARIMAX
        # SARIMA(0,1,1) with no seasonality for weekly data
        model = SARIMAX(y_train, order=(0,1,1), seasonal_order=(0,0,0,0))
        fitted = model.fit(disp=False, maxiter=200)
        forecast = fitted.forecast(steps=horizon)
        
        forecast_idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), 
                                     periods=horizon, freq="W-SUN")
        return pd.Series(np.maximum(0, forecast.values), index=forecast_idx), None
    except Exception as e:
        return None, str(e)

def forecast_hurdle(y_train, horizon):
    """Hurdle model - two-part: P(non-zero) × E(amount|non-zero)"""
    try:
        y = y_train.values.astype(float)
        
        # Part 1: Probability of non-zero
        non_zero_mask = y > 0
        p_nonzero = np.mean(non_zero_mask)
        
        # Part 2: Expected value given non-zero
        non_zero_values = y[non_zero_mask]
        
        if len(non_zero_values) == 0:
            return pd.Series([0] * horizon, 
                           index=pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), 
                                             periods=horizon, freq="W-SUN")), None
        
        # Use exponential smoothing for non-zero values
        if len(non_zero_values) >= 5:
            alpha = 0.3
            smoothed = non_zero_values[0]
            for val in non_zero_values[1:]:
                smoothed = alpha * val + (1 - alpha) * smoothed
            expected_nonzero = smoothed
        else:
            expected_nonzero = non_zero_values.mean()
        
        # Combine: P(Y>0) × E(Y|Y>0)
        forecast_value = p_nonzero * expected_nonzero
        
        forecast_idx = pd.date_range(y_train.index[-1] + pd.Timedelta(weeks=1), 
                                     periods=horizon, freq="W-SUN")
        return pd.Series([forecast_value] * horizon, index=forecast_idx), None
    except Exception as e:
        return None, str(e)

def select_best_forecast(y_historical, global_model=None, global_feature_cols=None, 
                        sku_id=None, sku_stats=None):
    """
    Test ALL 11+ models and select the best one based on wMAPE
    
    Models tested:
    1. Neural Network (Local)
    2. XGBoost (Local)
    3. Random Forest (Local)
    4. Holt Damped
    5. SARIMA(0,1,1)
    6. Hurdle
    7. TSB
    8. SBA
    9. Croston
    10. XGBoost (Global) - if available
    """
    
    # Calculate demand characteristics for classification
    adi, cv2, demand_class = calculate_demand_characteristics(y_historical)
    
    if adi is None or cv2 is None:
        print(f"  Demand Class: {demand_class} (ADI=NA, CV²=NA)")
    else:
        print(f"  Demand Class: {demand_class} (ADI={adi:.2f}, CV²={cv2:.2f})")
    
    # Split into train/validation
    val_size = min(12, len(y_historical) // 4)  # Last 12 weeks or 25% for validation
    y_train = y_historical[:-val_size]
    y_val = y_historical[-val_size:]
    val_horizon = len(y_val)
    
    print(f"  Training: {len(y_train)} weeks, Validation: {len(y_val)} weeks")
    
    # ALL MODELS - test everything
    methods = {
        # ML Models
        'Neural Network (Local)': lambda y, h: forecast_neural_network(y, h),
        'XGBoost (Local)': lambda y, h: forecast_xgboost(y, h),
        'Random Forest (Local)': lambda y, h: forecast_random_forest(y, h),
        
        # Time Series Models
        'Holt (Damped)': lambda y, h: forecast_holt_damped(y, h),
        'SARIMA(0,1,1)': lambda y, h: forecast_sarima(y, h),
        
        # Intermittent Demand Models
        'TSB': lambda y, h: forecast_tsb(y, h),
        'SBA': lambda y, h: forecast_sba(y, h),
        'Croston': lambda y, h: forecast_croston(y, h),
        'Hurdle': lambda y, h: forecast_hurdle(y, h),
    }
    
    # Add Global XGBoost if available
    if global_model is not None and sku_id is not None and sku_stats is not None:
        methods['XGBoost (Global)'] = lambda y, h: forecast_global_xgboost(
            global_model, global_feature_cols, y, sku_id, sku_stats, h
        )
    
    best_method = None
    best_wmape = float("inf")
    best_forecast = None
    candidates = []
    
    print("  Evaluating ALL forecast methods:")
    
    for method_name, method_func in methods.items():
        try:
            forecast_val, err = method_func(y_train, val_horizon)
            
            if forecast_val is None or err:
                print(f"    {method_name}: FAILED ({err})")
                continue
            
            # Calculate wMAPE on validation
            w = wmape(y_val.values, forecast_val.values)
            w_finite = np.isfinite(w)
            w_display = f"{w:.2f}%" if w_finite else "NA"
            
            # Calculate ratio to historical average for sanity check
            forecast_avg = forecast_val.mean()
            historical_avg = y_historical.mean()
            ratio = forecast_avg / historical_avg if historical_avg > 0 else 0
            
            print(f"    {method_name}: wMAPE={w_display}, Forecast avg={forecast_avg:.2f} (ratio={ratio:.2f}x)")

            if not w_finite:
                adjusted = float("inf")
                bias_note = "no signal"
            elif historical_avg > 0:
                if ratio < 0.5:
                    adjusted = w * 1.5
                    bias_note = "severe under-forecast"
                elif ratio < 0.8:
                    adjusted = w * 1.2
                    bias_note = "under-forecast"
                elif ratio > 2.0:
                    adjusted = w * 1.1
                    bias_note = "over-forecast"
                else:
                    adjusted = w
                    bias_note = "reasonable"
            else:
                adjusted = w
                bias_note = "no history"

            if w_finite:
                candidates.append({
                    "name": method_name,
                    "wmape": w,
                    "adjusted": adjusted,
                    "bias_note": bias_note,
                })
                
        except Exception as e:
            print(f"    {method_name}: ERROR ({e})")
            continue
    
    if candidates:
        best = min(candidates, key=lambda x: x["adjusted"])
        best_method = best["name"]
        best_wmape = best["wmape"]
    else:
        historical_avg = y_historical.mean()
        best_forecast = pd.Series(
            [historical_avg] * FUTURE_FORECAST_WEEKS,
            index=pd.date_range(y_historical.index[-1] + pd.Timedelta(weeks=1),
                                periods=FUTURE_FORECAST_WEEKS, freq="W-SUN"),
        )
        best_method = "Historical Avg (fallback)"
        best_wmape = float("nan")

    # Generate final forecast with best method
    if best_method and best_forecast is None:
        best_forecast, _ = methods[best_method](y_historical, FUTURE_FORECAST_WEEKS)
        
        # Sanity check: if wMAPE > 100%, all methods failed
        if best_wmape > 100:
            print(f"    ⚠ WARNING: All methods have wMAPE > 100% (all performing poorly)")
            print(f"    ⚠ Using historical average as fallback forecast")
            # Use historical average as fallback
            historical_avg = y_historical.mean()
            best_forecast = pd.Series([historical_avg] * FUTURE_FORECAST_WEEKS, 
                                     index=pd.date_range(y_historical.index[-1] + pd.Timedelta(weeks=1), 
                                                        periods=FUTURE_FORECAST_WEEKS, freq="W-SUN"))
            best_method = f"{best_method} (fallback to avg)"
    
    return best_method, best_wmape, best_forecast, adi, cv2, demand_class


# ============================================================
# GLOBAL XGBOOST MODEL - Train on ALL SKUs together
# ============================================================

def build_global_training_data(conn, sku_master, cutoff_date):
    """
    Load historical data for ALL SKUs and create a unified training dataset
    Each SKU uses data from its first sale month forward
    Returns: DataFrame with features for global model
    """
    print("\n" + "="*70)
    print("BUILDING GLOBAL MODEL DATASET")
    print(f"Processing SKUs with sales since {cutoff_date.date()}")
    print("Each SKU uses data from its first sale month forward")
    print("="*70)
    
    all_data = []
    sku_encodings = {}
    
    for idx, row in sku_master.iterrows():
        sku = row['ITEM_NUMBER']
        
        # Load sales from cutoff to find first sale
        sales_df_temp = load_sales_history(conn, sku, cutoff_date)
        first_sale_month = find_first_sale_month(sales_df_temp)
        
        if first_sale_month is None:
            continue
        
        # Load sales from first sale month forward
        sales_df = load_sales_history(conn, sku, first_sale_month)
        y_historical = build_daily_series(sales_df, first_sale_month)

        if y_historical is None or len(y_historical) < 100:
            continue

        # Cap extreme outliers before using for model training
        y_historical = cap_outliers(y_historical, n_mad=4.0)

        # Assign SKU encoding
        sku_encodings[sku] = len(sku_encodings)
        sku_id = sku_encodings[sku]
        
        # Calculate SKU-level statistics for features
        total_demand = y_historical.sum()
        avg_demand = y_historical.mean()
        std_demand = y_historical.std()
        zero_pct = (y_historical == 0).sum() / len(y_historical)
        
        # Create features for each time point
        for i in range(28, len(y_historical)):  # Need at least 28 days of lags
            record = {
                'sku_id': sku_id,
                'target': y_historical.iloc[i],
                'date': y_historical.index[i],
                # Lag features
                'lag_1': y_historical.iloc[i-1],
                'lag_7': y_historical.iloc[i-7],
                'lag_14': y_historical.iloc[i-14],
                'lag_28': y_historical.iloc[i-28],
                # Rolling averages
                'roll_7': y_historical.iloc[i-7:i].mean(),
                'roll_14': y_historical.iloc[i-14:i].mean(),
                'roll_28': y_historical.iloc[i-28:i].mean(),
                # Time features
                'dayofweek': y_historical.index[i].dayofweek,
                'day': y_historical.index[i].day,
                'month': y_historical.index[i].month,
                'quarter': y_historical.index[i].quarter,
                # SKU-level features
                'sku_avg_demand': avg_demand,
                'sku_std_demand': std_demand,
                'sku_zero_pct': zero_pct,
            }
            all_data.append(record)
        
        if (idx + 1) % 50 == 0:
            print(f"  Processed {idx + 1}/{len(sku_master)} SKUs...")
    
    df = pd.DataFrame(all_data)
    print(f"\n✓ Global dataset built: {len(df):,} records from {len(sku_encodings)} SKUs")
    
    return df, sku_encodings

def train_global_xgboost(df_global):
    """
    Train a single XGBoost model on all SKUs combined
    """
    print("\nTraining global XGBoost model...")
    
    # Features
    feature_cols = ['sku_id', 'lag_1', 'lag_7', 'lag_14', 'lag_28', 
                    'roll_7', 'roll_14', 'roll_28',
                    'dayofweek', 'day', 'month', 'quarter',
                    'sku_avg_demand', 'sku_std_demand', 'sku_zero_pct']
    
    X = df_global[feature_cols].values
    y = df_global['target'].values
    
    # Split into train/validation
    split_idx = int(len(X) * 0.8)
    X_train, X_val = X[:split_idx], X[split_idx:]
    y_train, y_val = y[:split_idx], y[split_idx:]
    
    # Train model
    model = XGBRegressor(
        n_estimators=500,
        learning_rate=0.05,
        max_depth=8,
        min_child_weight=3,
        subsample=0.8,
        colsample_bytree=0.8,
        random_state=42,
        n_jobs=-1
    )
    
    model.fit(X_train, y_train)
    
    # Evaluate
    y_pred = model.predict(X_val)
    val_wmape = wmape(y_val, y_pred)
    
    print(f"✓ Global model trained: wMAPE={val_wmape:.2f}%")
    
    return model, feature_cols

def forecast_global_xgboost(model, feature_cols, y_historical, sku_id, sku_stats, horizon):
    """
    Generate forecast for a specific SKU using the global model
    """
    try:
        preds = []
        hist = y_historical.copy()
        
        for i in range(horizon):
            future_date = hist.index[-1] + pd.Timedelta(days=1)
            
            # Build feature vector
            features = {
                'sku_id': sku_id,
                'lag_1': hist.iloc[-1],
                'lag_7': hist.iloc[-7],
                'lag_14': hist.iloc[-14],
                'lag_28': hist.iloc[-28] if len(hist) >= 28 else hist.iloc[0],
                'roll_7': hist.iloc[-7:].mean(),
                'roll_14': hist.iloc[-14:].mean(),
                'roll_28': hist.iloc[-28:].mean() if len(hist) >= 28 else hist.mean(),
                'dayofweek': future_date.dayofweek,
                'day': future_date.day,
                'month': future_date.month,
                'quarter': future_date.quarter,
                'sku_avg_demand': sku_stats['avg_demand'],
                'sku_std_demand': sku_stats['std_demand'],
                'sku_zero_pct': sku_stats['zero_pct'],
            }
            
            X_new = np.array([[features[col] for col in feature_cols]])
            pred = max(0.0, model.predict(X_new)[0])
            preds.append(pred)
            
            # Add prediction to history for next iteration
            new_row = pd.Series([pred], index=[future_date])
            hist = pd.concat([hist, new_row])
        
        idx = pd.date_range(y_historical.index[-1] + pd.Timedelta(days=1), periods=horizon, freq="D")
        return pd.Series(preds, index=idx, dtype=float), None
    except Exception as e:
        return None, str(e)

def calculate_inventory_metrics(y_historical_weekly, lead_time_days):
    """
    Calculate metrics using WEEKLY data
    
    Parameters:
    - y_historical_weekly: Weekly demand series
    - lead_time_days: Lead time in days
    
    Returns inventory metrics in SF units
    """
    # Total demand (in SF)
    total_demand = y_historical_weekly.sum()
    total_weeks = len(y_historical_weekly)
    total_months = total_weeks / WEEKS_PER_MONTH
    
    # Convert lead time to weeks
    lead_time_weeks = lead_time_days / DAYS_PER_WEEK
    
    # Weekly demand statistics
    weekly_mean_demand = total_demand / total_weeks if total_weeks > 0 else 0
    weekly_std_demand = y_historical_weekly.std()  # Actual std from weekly data
    
    # Monthly demand (for reporting)
    avg_monthly_demand = total_demand / total_months if total_months > 0 else 0
    
    # Daily equivalents (for display only)
    daily_mean_demand = weekly_mean_demand / DAYS_PER_WEEK
    
    # Lead time demand (in SF)
    lead_time_mean_demand = weekly_mean_demand * lead_time_weeks
    
    # Standard deviation over lead time
    # Formula: σ_LT = σ_weekly × √(LT_weeks)
    sd_demand_over_lt = weekly_std_demand * np.sqrt(lead_time_weeks)
    
    # Monthly standard deviation
    sd_demand_monthly = weekly_std_demand * np.sqrt(WEEKS_PER_MONTH)
    
    # Safety stock (in SF)
    safety_stock = Z_SCORE * sd_demand_over_lt
    
    # CAP SAFETY STOCK: Maximum 3 months of average demand
    max_safety_stock = avg_monthly_demand * 1.0
    if safety_stock > max_safety_stock:
        safety_stock = max_safety_stock
    
    # Reorder point (in SF) = Lead time demand + Safety stock
    reorder_point = lead_time_mean_demand + safety_stock
    
    # Reorder Quantity (in SF)
    # Cover lead time + 1 week review period
    review_period_weeks = 1  # Weekly review
    total_period_weeks = lead_time_weeks + review_period_weeks
    reorder_quantity = weekly_mean_demand * total_period_weeks
    
    return {
        'avg_monthly_demand': avg_monthly_demand,
        'avg_weekly_demand': weekly_mean_demand,
        'daily_mean_demand': daily_mean_demand,
        'weekly_std_demand': weekly_std_demand,
        'lead_time_days': lead_time_days,
        'lead_time_weeks': lead_time_weeks,
        'lead_time_mean_demand': lead_time_mean_demand,
        'sd_demand_over_lt': sd_demand_over_lt,
        'sd_demand_monthly': sd_demand_monthly,
        'safety_stock': safety_stock,
        'reorder_point': reorder_point,
        'reorder_quantity': reorder_quantity,
    }

def simulate_monthly_projection(inventory_position, weekly_forecast, reorder_point, reorder_qty,
                                sales_df, as_of_date, safety_stock, months_ahead=12):
    """
    Simulate monthly inventory starting at current month.

    - Historical rows are handled outside this function.
    - Adds a catchup row for the current month (MTD actual + forecast remainder).
    - Forecast rows cover the next `months_ahead` months.
    """
    if as_of_date is None:
        as_of_date = pd.Timestamp.now().normalize()
    else:
        as_of_date = pd.Timestamp(as_of_date).normalize()

    month_start = as_of_date.replace(day=1)
    month_end = (month_start + pd.offsets.MonthEnd(0)).normalize()
    as_of_end = as_of_date + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

    # MTD actuals for current month (only through today)
    sales_df = sales_df.copy()
    sales_df['SALES_DATE'] = pd.to_datetime(sales_df['SALES_DATE'], errors='coerce')
    sales_df = sales_df[sales_df['SALES_DATE'].notna()]
    sales_df = sales_df[sales_df['SALES_DATE'] <= as_of_end]
    mtd_mask = (sales_df['SALES_DATE'] >= month_start) & (sales_df['SALES_DATE'] <= as_of_end)
    mtd_actual = float(sales_df.loc[mtd_mask, 'QTY_SOLD_SF'].sum())

    remaining_days = max(0, (month_end - as_of_date).days)
    remainder_fc = float(weekly_forecast.mean()) * (remaining_days / DAYS_PER_WEEK)
    total_month_demand = mtd_actual + remainder_fc

    # Convert weekly forecast to daily
    daily_idx = pd.date_range(
        weekly_forecast.index[0],
        weekly_forecast.index[-1] + pd.Timedelta(days=6),
        freq='D'
    )
    daily_forecast = pd.Series(index=daily_idx, dtype=float)
    for week_end, week_demand in weekly_forecast.items():
        week_start = week_end - pd.Timedelta(days=6)
        mask = (daily_idx >= week_start) & (daily_idx <= week_end)
        daily_forecast.loc[mask] = week_demand / DAYS_PER_WEEK

    df_forecast = pd.DataFrame({'Date': daily_forecast.index, 'Demand': daily_forecast.values})
    df_forecast['Month'] = pd.to_datetime(df_forecast['Date']).dt.to_period('M').dt.to_timestamp()
    monthly_demand = df_forecast.groupby('Month')['Demand'].sum().reset_index()

    # Build forecast months starting next month
    start_fc_month = (month_start + pd.offsets.MonthBegin(1)).normalize()
    monthly_demand = monthly_demand[monthly_demand['Month'] >= start_fc_month]
    monthly_demand = monthly_demand.sort_values('Month').head(months_ahead)
    if len(monthly_demand) < months_ahead:
        avg_month_demand = float(weekly_forecast.mean()) * WEEKS_PER_MONTH if len(weekly_forecast) else 0.0
        last_month = monthly_demand['Month'].max() if not monthly_demand.empty else (start_fc_month - pd.offsets.MonthBegin(1))
        extra_months = pd.date_range(
            last_month + pd.offsets.MonthBegin(1),
            periods=months_ahead - len(monthly_demand),
            freq="MS",
        )
        extra = pd.DataFrame({"Month": extra_months, "Demand": avg_month_demand})
        monthly_demand = pd.concat([monthly_demand, extra], ignore_index=True)

    projections = []
    inventory_level = inventory_position

    # Catchup row for current month
    beginning_inv = inventory_level
    reorder_triggered = (beginning_inv - total_month_demand) < reorder_point
    if reorder_triggered:
        inventory_level += reorder_qty
    ending_inv = max(0, inventory_level - total_month_demand)
    projections.append({
        'Month': month_start,
        'Historical Demand': mtd_actual,
        'Safety Stock': safety_stock,
        'Beginning Inventory': beginning_inv,
        'Forecast': remainder_fc,
        'Order Quantity': reorder_qty if reorder_triggered else 0,
        'Ending Inventory': ending_inv,
        'Row_Type': 'CATCHUP'
    })
    inventory_level = ending_inv

    for _, row in monthly_demand.iterrows():
        month = row['Month']
        demand = row['Demand']

        beginning_inv = inventory_level
        reorder_triggered = beginning_inv <= reorder_point
        if reorder_triggered:
            inventory_level += reorder_qty
            reorder_qty_applied = reorder_qty
        else:
            reorder_qty_applied = 0

        ending_inv = max(0, inventory_level - demand)

        projections.append({
            'Month': month,
            'Historical Demand': np.nan,
            'Safety Stock': safety_stock,
            'Beginning Inventory': beginning_inv,
            'Forecast': demand,
            'Order Quantity': reorder_qty_applied,
            'Ending Inventory': ending_inv,
            'Row_Type': 'FCST'
        })

        inventory_level = ending_inv

    return pd.DataFrame(projections)

def find_first_sale_month(sales_df):
    """
    Find the first month with non-zero sales
    Returns: Timestamp of first day of that month, or None if no sales
    """
    if sales_df.empty:
        return None
    
    # Filter to only non-zero sales
    sales_with_qty = sales_df[sales_df['QTY_SOLD_SF'] > 0].copy()
    
    if sales_with_qty.empty:
        return None
    
    # Get the earliest sale date
    first_sale_date = sales_with_qty['SALES_DATE'].min()
    
    # Return the first day of that month
    return pd.Timestamp(year=first_sale_date.year, month=first_sale_date.month, day=1)

def process_sku(conn, sku_row, lead_times_excel, abc_map, global_model=None, global_feature_cols=None, 
                sku_encodings=None):
    """Process a single SKU - optionally using global model"""
    sku = sku_row['ITEM_NUMBER']
    print(f"\n{'='*70}")
    print(f"Processing: {sku}")
    print(f"{'='*70}")
    
    # Lead time
    if sku in lead_times_excel:
        lead_time_days = lead_times_excel[sku]
        lt_source = "Excel"
    else:
        lead_time_days = float(sku_row['LEAD_TIME_IMLT'])
        lt_source = "Gartman"
    
    # Step 1: Load sales from cutoff date to find first actual sale
    cutoff_date = pd.Timestamp(CUTOFF_DATE)
    sales_df_temp = load_sales_history(conn, sku, cutoff_date)
    
    # Find first sale month
    first_sale_month = find_first_sale_month(sales_df_temp)
    
    if first_sale_month is None:
        print(f"  SKIPPED: No sales found since {CUTOFF_DATE}")
        return None
    
    # Use first sale month as start date
    start_date = first_sale_month
    
    print(f"  First sale: {start_date.date()}")
    
    # Step 2: Load sales from first sale month forward and aggregate to WEEKLY
    sales_df = load_sales_history(conn, sku, start_date)
    y_historical_weekly = build_weekly_series(sales_df, start_date)

    if y_historical_weekly is None or len(y_historical_weekly) < 20:
        print(f"  SKIPPED: Insufficient data (need at least 20 weeks)")
        return None

    # Cap extreme outliers (e.g., data entry errors) before forecasting
    y_historical_weekly = cap_outliers(y_historical_weekly, n_mad=4.0)
    
    # Calculate statistics
    total_weeks = len(y_historical_weekly)
    total_demand = y_historical_weekly.sum()
    weekly_avg = y_historical_weekly.mean()
    daily_avg = weekly_avg / 7
    non_zero_weeks = (y_historical_weekly > 0).sum()
    non_zero_pct = 100 * non_zero_weeks / total_weeks
    
    print(f"  History: {total_weeks} weeks, Total: {total_demand:.2f} SF")
    print(f"  Weekly avg: {weekly_avg:.2f} SF/week, Daily avg: {daily_avg:.2f} SF/day")
    print(f"  Non-zero weeks: {non_zero_weeks} ({non_zero_pct:.1f}%)")
    
    # Prepare parameters for global model if available
    sku_id = None
    sku_stats = None
    if global_model is not None and sku_encodings is not None and sku in sku_encodings:
        sku_id = sku_encodings[sku]
        sku_stats = {
            'avg_demand': y_historical_weekly.mean(),
            'std_demand': y_historical_weekly.std(),
            'zero_pct': (y_historical_weekly == 0).sum() / len(y_historical_weekly)
        }
    
    # Select best forecast method - includes demand classification!
    best_method, best_wmape, weekly_forecast, adi, cv2, demand_class = select_best_forecast(
        y_historical_weekly, global_model, global_feature_cols, sku_id, sku_stats
    )
    
    if weekly_forecast is None:
        print(f"  SKIPPED: All methods failed")
        return None
    
    print(f"  Best Method: {best_method} (wMAPE: {_format_metric(best_wmape)}%)")
    
    # Check forecast sanity (compare weekly averages)
    forecast_weekly_avg = weekly_forecast.mean()
    historical_weekly_avg = y_historical_weekly.mean()
    forecast_ratio = forecast_weekly_avg / historical_weekly_avg if historical_weekly_avg > 0 else 0
    
    # Convert to daily for display
    forecast_daily_avg = forecast_weekly_avg / 7
    total_52_weeks = weekly_forecast.sum()
    
    print(f"  Future forecast: Weekly avg={forecast_weekly_avg:.2f} SF/week, Daily avg={forecast_daily_avg:.2f} SF/day")
    print(f"  Total 52 weeks={total_52_weeks:.2f} SF")
    
    # Warn if forecast is very different from history
    if forecast_ratio < 0.5 or forecast_ratio > 2.0:
        print(f"  ⚠ WARNING: Forecast ({forecast_weekly_avg:.2f}/wk) is {forecast_ratio:.2f}x historical ({historical_weekly_avg:.2f}/wk)")
        print(f"  ⚠ This may indicate poor model fit for highly intermittent demand")
    
    # Calculate metrics using WEEKLY data
    metrics = calculate_inventory_metrics(y_historical_weekly, lead_time_days)
    
    # ABC-BASED SAFETY STOCK TUNING
    sku_abc = abc_map.get(sku, "C")  # Default to C if not in map
    ss_base = float(metrics["safety_stock"])
    lead_time_weeks = float(metrics["lead_time_weeks"])
    
    # Only compute lumpy SS for ERRATIC/LUMPY demand
    if demand_class in ("LUMPY", "ERRATIC"):
        ss_lumpy = compute_ss_lumpy_from_percentile(y_historical_weekly, lead_time_weeks, sku_abc, ss_base)
    else:
        ss_lumpy = ss_base
    
    # Apply ABC-based uplift scalar
    ss_final_pre, uplift_raw, uplift_applied = apply_ss_uplift_scalar(ss_base, ss_lumpy, sku_abc)
    
    # Store diagnostics for Excel output
    metrics["sku_abc"] = sku_abc
    metrics["ss_base"] = ss_base
    metrics["ss_lumpy"] = ss_lumpy
    metrics["ss_uplift_raw"] = uplift_raw
    metrics["ss_uplift_applied_pre_budget"] = uplift_applied
    metrics["safety_stock"] = ss_final_pre
    metrics["reorder_point"] = metrics["lead_time_mean_demand"] + metrics["safety_stock"]
    
    if demand_class in ("LUMPY", "ERRATIC"):
        print(f"  SS tuning: ABC={sku_abc}, SS_base={ss_base:.2f}, SS_lumpy={ss_lumpy:.2f}, uplift_raw={uplift_raw:.2f}, uplift_applied={uplift_applied:.2f}")
    
    # Inventory position
    inv_position = float(sku_row['INVENTORY_POSITION'])
    
    # Monthly projection with catchup (current month) + next 12 months
    as_of_date = pd.Timestamp.now().normalize()
    monthly_proj = simulate_monthly_projection(
        inv_position,
        weekly_forecast,
        metrics['reorder_point'],
        metrics['reorder_quantity'],
        sales_df,
        as_of_date,
        metrics['safety_stock'],
        months_ahead=12
    )

    # Add historical monthly demand rows above catchup/forecast rows
    if not sales_df.empty:
        df_hist = sales_df.copy()
        df_hist['Month'] = pd.to_datetime(df_hist['SALES_DATE']).dt.to_period('M').dt.to_timestamp()
        df_hist = df_hist.groupby('Month', as_index=False)['QTY_SOLD_SF'].sum().sort_values('Month')
        df_hist = df_hist[df_hist['Month'] < as_of_date.replace(day=1)]
        if not df_hist.empty:
            hist_series = df_hist.set_index('Month')['QTY_SOLD_SF']
            hist_capped = cap_outliers(hist_series, n_mad=4.0)
            if not np.allclose(hist_series.values, hist_capped.values, equal_nan=True):
                print(f"  Historical monthly outliers capped for {sku}")
            df_hist['QTY_SOLD_SF'] = hist_capped.values
        hist_rows = []
        for _, r in df_hist.iterrows():
            hist_rows.append({
                'Month': r['Month'],
                'Historical Demand': r['QTY_SOLD_SF'],
                'Safety Stock': metrics['safety_stock'],
                'Beginning Inventory': np.nan,
                'Forecast': np.nan,
                'Order Quantity': np.nan,
                'Ending Inventory': np.nan,
                'Row_Type': 'HIST'
            })
        if hist_rows:
            monthly_proj = pd.concat([pd.DataFrame(hist_rows), monthly_proj], ignore_index=True)

    # Add vendor fields to monthly projections (will be added alongside SKU in final output)
    monthly_proj['vendor_number'] = sku_row['VENDOR_NUMBER'] if 'VENDOR_NUMBER' in sku_row else ''
    monthly_proj['vendor_name'] = sku_row['VENDOR_NAME'] if 'VENDOR_NAME' in sku_row else ''
    monthly_proj['collection'] = sku_row['COLLECTION'] if 'COLLECTION' in sku_row else ''
    monthly_proj['description'] = sku_row['DESCRIPTION'] if 'DESCRIPTION' in sku_row else ''
    
    # First reorder month
    first_reorder = monthly_proj[pd.to_numeric(monthly_proj['Order Quantity'], errors='coerce') > 0]
    first_reorder_month = first_reorder.iloc[0]['Month'] if len(first_reorder) > 0 else 'None'
    
    print(f"  ROP: {metrics['reorder_point']:.2f}, ROQ: {metrics['reorder_quantity']:.2f}")
    print(f"  First Reorder: {first_reorder_month}")
    
    result = {
        'sku': sku,
        'description': sku_row['DESCRIPTION'],
        'vendor_number': sku_row['VENDOR_NUMBER'] if 'VENDOR_NUMBER' in sku_row else '',
        'vendor_name': sku_row['VENDOR_NAME'] if 'VENDOR_NAME' in sku_row else '',
        'collection': sku_row['COLLECTION'] if 'COLLECTION' in sku_row else '',
        'available_sf': float(sku_row['AVAILABLE_SF']),
        'on_po_sf': float(sku_row['ON_PO_SF']),
        'backorder_sf': float(sku_row['BACKORDER_SF']),
        'inventory_position': inv_position,
        'lead_time_days': lead_time_days,
        'lead_time_source': lt_source,
        'demand_class': demand_class,
        'adi': adi,
        'cv2': cv2,
        'forecast_method': best_method,
        'forecast_wmape': best_wmape,
        'first_reorder_month': first_reorder_month,
        **metrics
    }
    
    return result, monthly_proj


def apply_global_inv_cap(df_results, df_monthly, inv_cap_sf):
    """
    Apply global ending inventory cap using min + flex scaling
    
    Strategy:
    1. Cap individual SKUs by ABC MOI limits first
    2. Calculate minimum safe levels (lead time demand)
    3. Calculate flex inventory (nice-to-have above minimum)
    4. Scale flex proportionally to fit within global cap
    
    Parameters:
    - df_results: DataFrame with inventory metrics per SKU
    - df_monthly: DataFrame with monthly projections per SKU
    - inv_cap_sf: Global inventory cap in SF
    
    Returns:
    - df_results: Updated with adjusted targets
    - df_monthly: Updated with recalculated projections
    """
    
    print(f"\n{'='*70}")
    print("APPLYING GLOBAL ENDING INVENTORY CAP")
    print(f"{'='*70}")
    print(f"  Global Inventory Cap: {inv_cap_sf:,.0f} SF")

    df_hist = pd.DataFrame()
    if 'Row_Type' in df_monthly.columns:
        df_hist = df_monthly[df_monthly['Row_Type'] == 'HIST'].copy()
        df_monthly = df_monthly[df_monthly['Row_Type'] != 'HIST'].copy()
    if 'Month' in df_monthly.columns:
        df_monthly['Month'] = pd.to_datetime(df_monthly['Month'], errors='coerce')
        df_monthly = df_monthly.dropna(subset=['Month'])
    if not df_hist.empty and 'Month' in df_hist.columns:
        df_hist['Month'] = pd.to_datetime(df_hist['Month'], errors='coerce')
    
    # Step 1: Cap individual SKUs by ABC MOI limits
    print(f"\n  Step 1: Applying ABC MOI caps...")
    for idx, row in df_results.iterrows():
        sku_abc = row.get('sku_abc', 'C')
        avg_monthly_demand = row['avg_monthly_demand']
        moi_cap = ABC_MOI_CAPS.get(sku_abc, 3.0)
        
        # Calculate max inventory position for this SKU
        max_inv_position = avg_monthly_demand * moi_cap
        
        # Cap reorder point, but never below lead-time demand
        current_rop = row['reorder_point']
        lead_time_demand = row['lead_time_mean_demand']
        if current_rop > max_inv_position:
            new_rop = max(lead_time_demand, max_inv_position)
            df_results.at[idx, 'reorder_point'] = new_rop
            # Safety stock = ROP - lead time demand (never negative)
            new_ss = max(0, new_rop - lead_time_demand)
            df_results.at[idx, 'safety_stock'] = new_ss
    
    # Recalculate monthly projections with MOI-capped values
    print(f"    Recalculating monthly projections after MOI caps...")
    updated_monthly_projs = []
    
    for sku in df_results['sku'].unique():
        sku_metrics = df_results[df_results['sku'] == sku].iloc[0]
        sku_monthly = df_monthly[df_monthly['SKU'] == sku].copy()
        
        if sku_monthly.empty:
            continue
        
        # Get updated values
        new_rop = sku_metrics['reorder_point']
        roq = sku_metrics['reorder_quantity']
        inv_pos = sku_metrics['inventory_position']
        
        # Recalculate projections
        inventory_level = inv_pos
        for month_idx, month_row in sku_monthly.iterrows():
            demand = month_row['Forecast']
            beginning_inv = inventory_level

            reorder_triggered = beginning_inv <= new_rop
            if reorder_triggered:
                inventory_level += roq
                reorder_qty = roq
            else:
                reorder_qty = 0

            sku_monthly.at[month_idx, 'Beginning Inventory'] = beginning_inv
            sku_monthly.at[month_idx, 'Order Quantity'] = reorder_qty
            sku_monthly.at[month_idx, 'Ending Inventory'] = max(0, inventory_level - demand)

            inventory_level = max(0, inventory_level - demand)
        
        updated_monthly_projs.append(sku_monthly)
    
    df_monthly = pd.concat(updated_monthly_projs, ignore_index=True) if updated_monthly_projs else df_monthly
    
    # Step 2: Calculate total ending inventory and check against cap
    print(f"\n  Step 2: Checking total ending inventory...")
    
    # Get December 2026 (last month) ending inventory for each SKU
    last_month = df_monthly['Month'].max()
    df_last_month = df_monthly[df_monthly['Month'] == last_month].copy()
    total_ending_inv = df_last_month['Ending Inventory'].sum()
    
    print(f"    Current total ending inventory ({last_month}): {total_ending_inv:,.0f} SF")
    
    if total_ending_inv <= inv_cap_sf:
        print(f"    ✓ Within capacity ({total_ending_inv:,.0f} SF <= {inv_cap_sf:,.0f} SF)")
        print(f"    No further adjustment needed")
        if not df_hist.empty:
            combined = []
            for sku in df_results['sku'].unique():
                hist_rows = df_hist[df_hist['SKU'] == sku]
                fcst_rows = df_monthly[df_monthly['SKU'] == sku]
                if not hist_rows.empty or not fcst_rows.empty:
                    combined.append(pd.concat([hist_rows, fcst_rows], ignore_index=True))
            df_monthly = pd.concat(combined, ignore_index=True) if combined else df_monthly
        return df_results, df_monthly
    
    print(f"    ⚠ EXCEEDS capacity by {total_ending_inv - inv_cap_sf:,.0f} SF")
    print(f"\n  Step 3: Applying min + flex scaling...")
    
    # Step 3: Define minimums and flex
    df_results['min_safe_level'] = df_results['lead_time_mean_demand']
    df_results['target_inv'] = df_results['reorder_point'] + (df_results['reorder_quantity'] * 0.5)
    df_results['flex_inv'] = (df_results['target_inv'] - df_results['min_safe_level']).clip(lower=0)
    
    sum_min = df_results['min_safe_level'].sum()
    sum_flex = df_results['flex_inv'].sum()
    
    print(f"    Sum of minimums: {sum_min:,.0f} SF")
    print(f"    Sum of flex: {sum_flex:,.0f} SF")
    print(f"    Total unconstrained: {sum_min + sum_flex:,.0f} SF")
    
    if sum_min >= inv_cap_sf:
        print(f"    ⚠ WARNING: Minimums alone exceed capacity!")
        print(f"    Consider reducing lead times or raising capacity")
        # Still apply scaling, but with warning
        capacity_for_flex = inv_cap_sf * 0.5  # Allow 50% for minimums
    else:
        capacity_for_flex = inv_cap_sf - sum_min
    
    # Step 4: Scale flex to fit
    if sum_flex > capacity_for_flex:
        scale = capacity_for_flex / sum_flex
        print(f"    Scaling factor: {scale:.3f}")
        df_results['new_target_inv'] = df_results['min_safe_level'] + (df_results['flex_inv'] * scale)
    else:
        df_results['new_target_inv'] = df_results['target_inv']
        scale = 1.0
    
    # Step 5: Update reorder points to match new targets
    print(f"\n  Step 4: Updating reorder points and safety stocks...")
    for idx, row in df_results.iterrows():
        new_target = row['new_target_inv']
        lead_time_demand = row['lead_time_mean_demand']
        
        # New ROP is the new target minus half a reorder quantity
        new_rop = max(lead_time_demand, new_target - (row['reorder_quantity'] * 0.5))
        new_ss = max(0, new_rop - lead_time_demand)
        
        df_results.at[idx, 'reorder_point'] = new_rop
        df_results.at[idx, 'safety_stock'] = new_ss
    
    # Step 6: Recalculate monthly projections with scaled values
    print(f"    Recalculating monthly projections with scaled targets...")
    final_monthly_projs = []
    
    for sku in df_results['sku'].unique():
        sku_metrics = df_results[df_results['sku'] == sku].iloc[0]
        sku_monthly = df_monthly[df_monthly['SKU'] == sku].copy()
        
        if sku_monthly.empty:
            continue
        
        # Get updated values
        new_rop = sku_metrics['reorder_point']
        roq = sku_metrics['reorder_quantity']
        inv_pos = sku_metrics['inventory_position']
        
        # Recalculate projections
        inventory_level = inv_pos
        for month_idx, month_row in sku_monthly.iterrows():
            demand = month_row['Forecast']
            beginning_inv = inventory_level

            reorder_triggered = beginning_inv <= new_rop
            if reorder_triggered:
                inventory_level += roq
                reorder_qty = roq
            else:
                reorder_qty = 0

            sku_monthly.at[month_idx, 'Beginning Inventory'] = beginning_inv
            sku_monthly.at[month_idx, 'Order Quantity'] = reorder_qty
            sku_monthly.at[month_idx, 'Ending Inventory'] = max(0, inventory_level - demand)

            inventory_level = max(0, inventory_level - demand)
        
        final_monthly_projs.append(sku_monthly)
    
    df_monthly = pd.concat(final_monthly_projs, ignore_index=True) if final_monthly_projs else df_monthly
    
    # Verify final total
    df_last_month_final = df_monthly[df_monthly['Month'] == last_month].copy()
    final_total = df_last_month_final['Ending Inventory'].sum()
    
    print(f"\n  ✓ Final total ending inventory ({last_month}): {final_total:,.0f} SF")
    print(f"    Reduction: {total_ending_inv - final_total:,.0f} SF ({100*(total_ending_inv - final_total)/total_ending_inv:.1f}%)")
    print(f"    Within cap: {'YES' if final_total <= inv_cap_sf else 'NO (check minimums)'}")
    print(f"{'='*70}")

    if not df_hist.empty:
        combined = []
        for sku in df_results['sku'].unique():
            hist_rows = df_hist[df_hist['SKU'] == sku]
            fcst_rows = df_monthly[df_monthly['SKU'] == sku]
            if not hist_rows.empty or not fcst_rows.empty:
                combined.append(pd.concat([hist_rows, fcst_rows], ignore_index=True))
        df_monthly = pd.concat(combined, ignore_index=True) if combined else df_monthly

    return df_results, df_monthly

def run_inventory_planning():
    output_dir = Path(__file__).parent
    lead_times_excel, sf_per_pallet_map, pallets_per_container_map = load_lead_times(output_dir / LEADTIMES_XLSX)
    
    # Load forecast SKU list if enabled
    forecast_sku_list = None
    if USE_SKU_LIST and not FOCUS_SKU:
        forecast_sku_list = load_forecast_sku_list(output_dir / FORECAST_SKU_LIST_FILE)
    
    conn = _connect()
    single_sku = FOCUS_SKU.strip().upper() if FOCUS_SKU else None
    sku_master = get_active_skus(conn, single_sku, forecast_sku_list)
    
    if sku_master.empty:
        print("ERROR: No active SKUs found")
        conn.close()
        return
    
    # Mode description
    if single_sku:
        mode_desc = f"Single SKU: {single_sku}"
    elif USE_SKU_LIST and forecast_sku_list and len(forecast_sku_list) > 0:
        mode_desc = f"SKU List: {len(sku_master)} SKUs from {FORECAST_SKU_LIST_FILE}"
    else:
        mode_desc = f"All Active SKUs: {len(sku_master)} SKUs"
    
    model_desc = "Global XGBoost Model" if USE_GLOBAL_MODEL and not single_sku else "Per-SKU Models"
    print(f"\nMODE: {mode_desc}")
    print(f"FORECAST MODEL: {model_desc}")
    
    # Build ABC segmentation based on trailing 12 month SF volume
    print(f"\n{'='*70}")
    print("BUILDING ABC SEGMENTATION")
    print(f"{'='*70}")
    sku_list_for_volume = sku_master["ITEM_NUMBER"].astype(str).str.upper().tolist()
    vol12_df = load_trailing_12m_volume(conn, sku_list_for_volume)
    abc_map = build_abc_map(vol12_df)
    
    # Calculate total trailing 12m volume for global uplift budget
    total_vol_12m = float(vol12_df["VOL_12M"].sum()) if not vol12_df.empty else 0.0
    uplift_budget = GLOBAL_UPLIFT_BUDGET_PCT * total_vol_12m if ENABLE_GLOBAL_UPLIFT_BUDGET else None
    
    if ENABLE_GLOBAL_UPLIFT_BUDGET and uplift_budget:
        print(f"\n  Global Uplift Budget: {uplift_budget:,.2f} SF ({GLOBAL_UPLIFT_BUDGET_PCT*100:.1f}% of trailing 12m volume)")
    
    print(f"{'='*70}")
    
    # Train global model if enabled and processing multiple SKUs
    global_model = None
    global_feature_cols = None
    sku_encodings = None
    
    if USE_GLOBAL_MODEL and not single_sku:
        cutoff_date = pd.Timestamp(CUTOFF_DATE)
        df_global, sku_encodings = build_global_training_data(conn, sku_master, cutoff_date)
        
        if len(df_global) > 1000:  # Need sufficient data
            global_model, global_feature_cols = train_global_xgboost(df_global)
        else:
            print("⚠ Insufficient data for global model, falling back to per-SKU models")
    
    results = []
    all_monthly_projs = []
    
    for idx, row in sku_master.iterrows():
        try:
            result = process_sku(conn, row, lead_times_excel, abc_map, 
                               global_model, global_feature_cols, sku_encodings)
            if result:
                metrics, monthly_proj = result
                results.append(metrics)
                monthly_proj['SKU'] = metrics['sku']
                all_monthly_projs.append(monthly_proj)
        except Exception as e:
            print(f"  ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
    
    if not results:
        print("\nNo SKUs processed")
        if conn is not None:
            conn.close()
        return
    
    df_results = pd.DataFrame(results)
    df_monthly = pd.concat(all_monthly_projs, ignore_index=True) if all_monthly_projs else pd.DataFrame()

    if not df_results.empty and "sku" in df_results.columns:
        df_results["sf_per_pallet"] = df_results["sku"].map(sf_per_pallet_map)
        df_results["pallets_per_container"] = df_results["sku"].map(pallets_per_container_map)
    if not df_monthly.empty and "SKU" in df_monthly.columns:
        df_monthly["sf_per_pallet"] = df_monthly["SKU"].map(sf_per_pallet_map)
        df_monthly["pallets_per_container"] = df_monthly["SKU"].map(pallets_per_container_map)
    
    # APPLY GLOBAL UPLIFT BUDGET if enabled
    if ENABLE_GLOBAL_UPLIFT_BUDGET and uplift_budget is not None and uplift_budget > 0 and not df_results.empty:
        total_uplift = float(df_results["ss_uplift_applied_pre_budget"].sum())
        if total_uplift > uplift_budget:
            k = uplift_budget / total_uplift
            print(f"\n{'='*70}")
            print(f"⚠ GLOBAL UPLIFT BUDGET ADJUSTMENT")
            print(f"{'='*70}")
            print(f"  Total uplift={total_uplift:,.2f} SF")
            print(f"  Budget={uplift_budget:,.2f} SF")
            print(f"  Scaling factor k={k:.3f}")
            print(f"  All uplifts will be scaled down proportionally")
            
            # Scale only the uplift, keep base intact
            df_results["ss_uplift_applied"] = df_results["ss_uplift_applied_pre_budget"] * k
            df_results["safety_stock"] = df_results["ss_base"] + df_results["ss_uplift_applied"]
            df_results["reorder_point"] = df_results["lead_time_mean_demand"] + df_results["safety_stock"]
            
            print(f"  ✓ Adjusted safety stock for all SKUs")
            print(f"{'='*70}")
        else:
            # No adjustment needed
            df_results["ss_uplift_applied"] = df_results["ss_uplift_applied_pre_budget"]
            print(f"\n  ✓ Global uplift budget NOT exceeded (Total={total_uplift:,.2f} SF, Budget={uplift_budget:,.2f} SF)")
    else:
        # No budget enabled, just copy pre-budget values
        if not df_results.empty and "ss_uplift_applied_pre_budget" in df_results.columns:
            df_results["ss_uplift_applied"] = df_results["ss_uplift_applied_pre_budget"]
    
    # APPLY GLOBAL ENDING INVENTORY CAP if enabled
    if ENABLE_GLOBAL_INV_CAP and not df_results.empty and not df_monthly.empty:
        df_results, df_monthly = apply_global_inv_cap(df_results, df_monthly, GLOBAL_INV_CAP_SF)
    
    # Determine output filename
    if single_sku:
        base_filename = 'inventory_plan_single.xlsx'
    elif USE_SKU_LIST and forecast_sku_list and len(forecast_sku_list) > 0:
        base_filename = 'inventory_plan_list.xlsx'
    else:
        base_filename = 'inventory_plan_all.xlsx'
    
    output_file = output_dir / base_filename
    
    # Try writing to the file
    saved = False
    for attempt in range(2):
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df_results.to_excel(writer, sheet_name='Inventory_Metrics', index=False)
                if not df_monthly.empty:
                    df_monthly_out = df_monthly.copy()
                    if 'Month' in df_monthly_out.columns:
                        df_monthly_out['Month'] = pd.to_datetime(
                            df_monthly_out['Month'], errors='coerce'
                        ).dt.strftime('%m/%d/%Y')
                    if 'Row_Type' in df_monthly_out.columns:
                        df_monthly_out = df_monthly_out.drop(columns=['Row_Type'])

                    # Align column order with sundries output
                    preferred = [
                        'Month',
                        'Historical Demand',
                        'Safety Stock',
                        'Beginning Inventory',
                        'Forecast',
                        'Order Quantity',
                        'Ending Inventory',
                        'SKU',
                        'vendor_number',
                        'collection',
                    ]
                    cols = [c for c in preferred if c in df_monthly_out.columns]
                    remaining = [c for c in df_monthly_out.columns if c not in cols]
                    if cols:
                        df_monthly_out = df_monthly_out[cols + remaining]
                    df_monthly_out.to_excel(writer, sheet_name='Monthly_Projections', index=False)
            saved = True
            print(f"\n✓ Saved: {output_file}")
            print(f"  Processed: {len(results)} SKUs")
            
            # Show method selection summary
            if len(df_results) > 0:
                print(f"\n{'='*70}")
                print("DEMAND CLASSIFICATION SUMMARY")
                print(f"{'='*70}")
                demand_counts = df_results['demand_class'].value_counts()
                for demand_class, count in demand_counts.items():
                    pct = 100 * count / len(df_results)
                    avg_adi = df_results[df_results['demand_class'] == demand_class]['adi'].mean()
                    avg_cv2 = df_results[df_results['demand_class'] == demand_class]['cv2'].mean()
                    print(
                        f"  {demand_class}: {count} SKUs ({pct:.1f}%) - "
                        f"Avg ADI={_format_metric(avg_adi)}, Avg CV2={_format_metric(avg_cv2)}"
                    )
                
                print(f"\n{'='*70}")
                print("FORECAST METHOD SELECTION SUMMARY")
                print(f"{'='*70}")
                method_counts = df_results['forecast_method'].value_counts()
                for method, count in method_counts.items():
                    pct = 100 * count / len(df_results)
                    avg_wmape = df_results[df_results['forecast_method'] == method]['forecast_wmape'].mean()
                    print(f"  {method}: {count} SKUs ({pct:.1f}%) - Avg wMAPE: {_format_metric(avg_wmape)}%")
                print(f"{'='*70}")
            
            break
        except PermissionError:
            if attempt == 0:
                # First attempt failed - try with timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                timestamped_filename = base_filename.replace('.xlsx', f'_{timestamp}.xlsx')
                output_file = output_dir / timestamped_filename
                print(f"\n⚠ Original file is locked. Trying: {timestamped_filename}")
            else:
                print(f"\n✗ ERROR: Could not save file. Please close Excel and try again.")
                print(f"  File path: {output_file}")
                break
    
    arrivals_df = pd.DataFrame()
    if conn is not None and not df_results.empty and "sku" in df_results.columns:
        arrivals_df = load_arrivals(conn, df_results["sku"].astype(str).unique().tolist())
        if not arrivals_df.empty:
            lt_map = df_results[["sku", "lead_time_days"]].rename(columns={"sku": "ITEM_NUMBER"})
            arrivals_df = arrivals_df.merge(lt_map, on="ITEM_NUMBER", how="left")
    if conn is not None:
        conn.close()
    payload = _build_webapp_payload(df_results, df_monthly, arrivals_df)
    _write_webapp_json(payload, WEBAPP_JSON_PATH)

    if not saved:
        print(f"\n⚠ WARNING: Results were not saved to disk!")
        print(f"  Please close the Excel file and run again.")

def _format_currency(value: float, digits: int = 2) -> str:
    if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
        return "--"
    return f"${value:,.{digits}f}"

def _df_to_records(df: pd.DataFrame) -> List[Dict]:
    if df is None or df.empty:
        return []
    df_out = df.copy()
    for col in df_out.columns:
        if pd.api.types.is_datetime64_any_dtype(df_out[col]):
            df_out[col] = df_out[col].dt.strftime("%Y-%m-%d")
    df_out = df_out.replace({pd.NA: None, np.nan: None, np.inf: None, -np.inf: None})
    return df_out.to_dict(orient="records")

def _format_number(value: float, digits: int = 2) -> str:
    if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
        return "--"
    return f"{value:,.{digits}f}"

def _format_metric(value: float, digits: int = 2) -> str:
    if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
        return "NA"
    return f"{value:.{digits}f}"

def _filter_forecast_months(series: pd.Series) -> pd.Series:
    if series is None or series.empty:
        return series
    if not isinstance(series.index, pd.DatetimeIndex):
        return series
    current_month_start = pd.Timestamp.today().to_period("M").to_timestamp()
    next_month_start = current_month_start + pd.offsets.MonthBegin(1)
    end_exclusive = current_month_start + pd.DateOffset(months=12)
    return series[(series.index >= next_month_start) & (series.index < end_exclusive)]

def _normalize_arrival_date(value: Any) -> pd.Timestamp:
    try:
        dt = pd.to_datetime(value, errors="coerce")
    except Exception:
        return pd.NaT
    if pd.isna(dt):
        return pd.NaT
    start = pd.Timestamp("2025-01-01")
    end = pd.Timestamp.today().normalize() + pd.Timedelta(days=730)
    if dt < start or dt > end:
        return pd.NaT
    return dt

def _format_arrival_date(value: Any) -> str:
    dt = _normalize_arrival_date(value)
    if pd.isna(dt):
        return "No Date"
    return dt.strftime("%m/%d/%Y")

def _format_arrival_value(value: Any, label: str = "No Date") -> str:
    dt = _normalize_arrival_date(value)
    if pd.isna(dt):
        return label
    return dt.strftime("%m/%d/%Y")

def _add_days_safe(value: Any, days: float) -> pd.Timestamp:
    base = _normalize_arrival_date(value)
    if pd.isna(base):
        return pd.NaT
    try:
        return base + pd.Timedelta(days=float(days))
    except Exception:
        return pd.NaT

def _mask_suffix(value: str) -> str:
    if not value:
        return "000"
    digits = "".join(ch for ch in value if ch.isdigit())
    if digits:
        return digits[-3:].zfill(3)
    return f"{abs(hash(value)) % 1000:03d}"

def _get_first_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for col in candidates:
        if col in df.columns:
            return col
    return None

def _demo_webapp_payload() -> Dict:
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return {
        "generated_at": datetime.now().isoformat(),
        "run_meta": {
            "title": "OMP Purchasing Dashboard",
            "run_timestamp_local": run_ts,
            "forecast_horizon_days": 30,
            "source": "Lead Times.xlsx:Final",
        },
        "items": [
            {
                "item_number": "GFAL09503",
                "description": "EURO OAK 9.5\" DOMA",
                "vendor_name": "FLO.I.T S.R.L",
                "vendor_number": "569",
                "available": 59731.11,
                "on_po": 0.0,
                "backorder": 0.0,
                "inventory_position": 59731.11,
                "lead_time_days": 53,
                "lead_time_source": "Excel",
                "lt_demand": 370.03,
                "suggested_order_qty": None,
                "forecast_series": {
                    "x": ["Nov 23", "Nov 30", "Dec 07", "Dec 14", "Dec 21"],
                    "y": [72.4, 74.8, 74.9, 74.9, 74.1],
                },
                "segmentation": {
                    "maturity": "Legacy",
                    "demand_pattern": "Lumpy",
                    "adi": 1.63,
                    "cv2": 1.48,
                    "first_sales_date": "2023-08-09",
                    "history_days": 862,
                },
                "daily_avg_demand": 12.33,
                "days_of_cover": 4842.7,
            }
        ],
        "arrivals": [
            {
                "PO_NUMBER": "450123",
                "ITEM_NUMBER": "GFAL09503",
                "DESCRIPTION": "EURO OAK 9.5\" DOMA",
                "QUANTITY_SF": 1200.0,
                "VENDOR_NAME": "FLO.I.T S.R.L",
                "VENDOR_NUMBER": "569",
                "COLLECTION": "ALLORA",
                "CONTAINER": "CONT-7742",
                "DUE_PORT": "2025-02-10",
                "DUE_INV": "2025-03-05",
            }
        ],
        "queue": [
            {
                "item_number": "GFAL09503",
                "description": "EURO OAK 9.5\" DOMA",
                "vendor_name": "FLO.I.T S.R.L",
                "vendor_number": "569",
                "available": 59731.11,
                "on_po": 0.0,
                "backorder": 0.0,
                "inventory_position": 59731.11,
                "lead_time_days": 53,
                "lt_demand": 370.03,
            }
        ],
        "sql_sources": [
            "gartman_sales_history_daily.sql",
        ],
    }

def _build_webapp_payload(df_results: pd.DataFrame, df_monthly: pd.DataFrame, df_arrivals: Optional[pd.DataFrame] = None) -> Dict:
    payload = _demo_webapp_payload()
    if df_results is None or df_results.empty:
        return payload

    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    payload["run_meta"] = {
        "title": "OMP Purchasing Dashboard",
        "run_timestamp_local": run_ts,
        "forecast_horizon_days": FUTURE_FORECAST_DAYS,
        "source": f"{LEADTIMES_XLSX}:Final",
    }

    col_month = _get_first_column(df_monthly, ["Month", "month"]) if df_monthly is not None else None
    col_forecast = _get_first_column(df_monthly, ["Forecast", "forecast"]) if df_monthly is not None else None
    col_hist = _get_first_column(df_monthly, ["Historical Demand", "Historical", "history"]) if df_monthly is not None else None
    col_sku = _get_first_column(df_monthly, ["SKU", "sku", "Item Number", "ITEM_NUMBER"]) if df_monthly is not None else None

    sku_series: Dict[str, Dict[str, List[float]]] = {}
    sku_first_dates: Dict[str, str] = {}
    sku_history_days: Dict[str, int] = {}

    row_type_col = _get_first_column(df_monthly, ["Row_Type", "row_type"]) if df_monthly is not None else None

    if df_monthly is not None and not df_monthly.empty and col_month and col_sku and (col_forecast or col_hist):
        df_m = df_monthly.copy()
        df_m[col_month] = pd.to_datetime(df_m[col_month], errors="coerce")
        df_m = df_m[df_m[col_month].notna()]
        for sku, sku_df in df_m.groupby(col_sku):
            value_col = col_forecast if col_forecast in sku_df.columns else col_hist
            series = sku_df.groupby(col_month)[value_col].sum().sort_index()
            if series.empty:
                continue
            if value_col == col_forecast:
                series = _filter_forecast_months(series)
                if series.empty:
                    continue
            last = series.tail(6)
            series_payload = {
                "fc_x": [d.strftime("%Y-%m-%d") for d in series.index],
                "fc_y": [float(v) for v in series.values],
                "x": [d.strftime("%Y-%m-%d") for d in series.index],
                "y": [float(v) for v in series.values],
            }
            if row_type_col and row_type_col in sku_df.columns:
                hist_mask = sku_df[row_type_col].astype(str).str.upper().eq("HIST")
                fc_mask = sku_df[row_type_col].astype(str).str.upper().isin(["FCST", "CATCHUP"])
                hist_series = sku_df.loc[hist_mask].groupby(col_month)[col_hist].sum().sort_index() if col_hist else None
                fc_series = sku_df.loc[fc_mask].groupby(col_month)[col_forecast].sum().sort_index() if col_forecast else None
                if fc_series is not None:
                    fc_series = _filter_forecast_months(fc_series)
                if hist_series is not None and not hist_series.empty:
                    series_payload["hist_x"] = [d.strftime("%Y-%m-%d") for d in hist_series.index]
                    series_payload["hist_y"] = [float(v) for v in hist_series.values]
                if fc_series is not None and not fc_series.empty:
                    series_payload["fc_x"] = [d.strftime("%Y-%m-%d") for d in fc_series.index]
                    series_payload["fc_y"] = [float(v) for v in fc_series.values]
                    series_payload["x"] = [d.strftime("%Y-%m-%d") for d in fc_series.index]
                    series_payload["y"] = [float(v) for v in fc_series.values]
            elif col_hist and col_hist in sku_df.columns:
                hist_series = sku_df[sku_df[col_hist].notna()].groupby(col_month)[col_hist].sum().sort_index()
                if not hist_series.empty:
                    series_payload["hist_x"] = [d.strftime("%Y-%m-%d") for d in hist_series.index]
                    series_payload["hist_y"] = [float(v) for v in hist_series.values]
            sku_series[str(sku)] = series_payload
            # Use historical series for first_sales_date if available, otherwise fall back to filtered series
            if hist_series is not None and not hist_series.empty:
                first_date = hist_series.index.min()
                last_date = hist_series.index.max()
            else:
                first_date = series.index.min() if not series.empty else pd.NaT
                last_date = series.index.max() if not series.empty else pd.NaT
            if pd.notna(first_date) and pd.notna(last_date):
                sku_first_dates[str(sku)] = first_date.strftime("%Y-%m-%d")
                sku_history_days[str(sku)] = max(0, int((last_date - first_date).days))

    items: List[Dict] = []
    for _, row in df_results.iterrows():
        sku = str(row.get("sku", ""))
        daily_avg = float(row.get("daily_mean_demand", 0.0) or 0.0)
        inv_pos = float(row.get("inventory_position", 0.0) or 0.0)
        days_cover = inv_pos / daily_avg if daily_avg > 0 else 0.0
        segmentation = {
            "maturity": str(row.get("sku_abc", "")) or "Legacy",
            "demand_pattern": str(row.get("demand_class", "")),
            "adi": float(row.get("adi", 0.0) or 0.0),
            "cv2": float(row.get("cv2", 0.0) or 0.0),
            "first_sales_date": sku_first_dates.get(sku),
            "history_days": sku_history_days.get(sku, 0),
        }
        items.append(
            {
                "item_number": sku,
                "description": str(row.get("description", "")),
                "vendor_name": str(row.get("vendor_name", "")) or str(row.get("collection", "")),
                "vendor_number": str(row.get("vendor_number", "")),
                "collection": str(row.get("collection", "")),
                "available": float(row.get("available_sf", 0.0) or 0.0),
                "on_po": float(row.get("on_po_sf", 0.0) or 0.0),
                "backorder": float(row.get("backorder_sf", 0.0) or 0.0),
                "inventory_position": inv_pos,
                "lead_time_days": int(row.get("lead_time_days", 0) or 0),
                "lead_time_source": str(row.get("lead_time_source", "")),
                "lt_demand": float(row.get("lead_time_mean_demand", 0.0) or 0.0),
                "suggested_order_qty": float(row.get("reorder_quantity", 0.0) or 0.0),
                "forecast_series": sku_series.get(sku, {}),
                "segmentation": segmentation,
                "daily_avg_demand": daily_avg,
                "days_of_cover": float(days_cover),
            }
        )

    df_queue = df_results.copy()
    if "reorder_point" in df_queue.columns and "inventory_position" in df_queue.columns:
        df_queue["risk_gap"] = df_queue["inventory_position"] - df_queue["reorder_point"]
        df_queue = df_queue.sort_values("risk_gap")
    queue_rows: List[Dict] = []
    for _, row in df_queue.head(8).iterrows():
        queue_rows.append(
            {
                "item_number": str(row.get("sku", "")),
                "description": str(row.get("description", "")),
                "vendor_name": str(row.get("vendor_name", "")) or str(row.get("collection", "")),
                "vendor_number": str(row.get("vendor_number", "")),
                "collection": str(row.get("collection", "")),
                "available": float(row.get("available_sf", 0.0) or 0.0),
                "on_po": float(row.get("on_po_sf", 0.0) or 0.0),
                "backorder": float(row.get("backorder_sf", 0.0) or 0.0),
                "inventory_position": float(row.get("inventory_position", 0.0) or 0.0),
                "lead_time_days": int(row.get("lead_time_days", 0) or 0),
                "lt_demand": float(row.get("lead_time_mean_demand", 0.0) or 0.0),
            }
        )

    payload["items"] = items
    payload["queue"] = queue_rows
    payload["Inventory_Metrics"] = _df_to_records(df_results)
    payload["Monthly_Projections"] = _df_to_records(df_monthly)
    if df_arrivals is not None and not df_arrivals.empty:
        payload["arrivals"] = _df_to_records(df_arrivals)
    else:
        payload["arrivals"] = []
    payload["generated_at"] = datetime.now().isoformat()
    return payload

def _write_webapp_json(payload: Dict, output_path: Path) -> None:
    def _sanitize_jsonable(value: Any) -> Any:
        if isinstance(value, float):
            if math.isnan(value) or math.isinf(value):
                return None
            return value
        if isinstance(value, (pd.Timestamp, datetime, date, np.datetime64)):
            if pd.isna(value):
                return None
            try:
                return pd.to_datetime(value).strftime("%Y-%m-%d")
            except Exception:
                return None
        if value is pd.NaT:
            return None
        if isinstance(value, dict):
            return {k: _sanitize_jsonable(v) for k, v in value.items()}
        if isinstance(value, list):
            return [_sanitize_jsonable(v) for v in value]
        return value

    try:
        safe_payload = _sanitize_jsonable(payload)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(safe_payload, f, indent=2, allow_nan=False)
        print(f"  Wrote webapp JSON: {output_path}")
    except Exception as exc:
        print(f"  WARNING: Failed to write webapp JSON ({exc})")

def _load_webapp_payload() -> Dict:
    if WEBAPP_JSON_PATH.exists():
        try:
            with open(WEBAPP_JSON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return _demo_webapp_payload()
    return _demo_webapp_payload()

def _load_sundries_payload() -> Dict:
    if SUNDRIES_JSON_PATH.exists():
        try:
            with open(SUNDRIES_JSON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return _demo_webapp_payload()
    return _demo_webapp_payload()

def _load_strip_sku_list() -> set:
    """Load list of SKUs for Veronica tab from StripSKUList.xlsx"""
    strip_path = Path(__file__).resolve().parent / STRIP_SKU_LIST_FILE
    if not strip_path.exists():
        return set()
    try:
        # SKUs are in the "Who Produces" sheet, first column, starting at row 2 (skip header)
        df = pd.read_excel(strip_path, sheet_name="Who Produces", header=0)
        if df.shape[0] == 0:
            return set()
        sku_col = df.columns[0]  # "Item Number" column
        df[sku_col] = df[sku_col].astype(str).str.strip().str.upper()
        sku_list = set(df[sku_col].dropna())
        sku_list = {sku for sku in sku_list if sku and sku != 'NAN' and len(sku) > 0}
        return sku_list
    except Exception:
        return set()

def _load_strip_sku_details() -> Dict[str, Dict]:
    """Load SKU details (Thickness, Width, Species, Grade/Cut, Edge, Bundles) from StripSKUList.xlsx"""
    strip_path = Path(__file__).resolve().parent / STRIP_SKU_LIST_FILE
    if not strip_path.exists():
        return {}
    try:
        df = pd.read_excel(strip_path, sheet_name="Who Produces", header=0)
        if df.shape[0] == 0:
            return {}
        # Normalize the Item Number column
        df["Item Number"] = df["Item Number"].astype(str).str.strip().str.upper()
        # Build a dict mapping SKU -> details
        details = {}
        for _, row in df.iterrows():
            sku = row["Item Number"]
            if sku and sku != 'NAN':
                details[sku] = {
                    "Description": str(row.get("Description", "")) if pd.notna(row.get("Description")) else "",
                    "Thickness": str(row.get("Thickness", "")) if pd.notna(row.get("Thickness")) else "",
                    "Width": str(row.get("Width", "")) if pd.notna(row.get("Width")) else "",
                    "Species": str(row.get("Species", "")) if pd.notna(row.get("Species")) else "",
                    "Grade/Cut": str(row.get("Grade/Cut", "")) if pd.notna(row.get("Grade/Cut")) else "",
                    "Edge": str(row.get("Edge", "")) if pd.notna(row.get("Edge")) else "",
                    "Bundles": str(row.get("Bundles", "")) if pd.notna(row.get("Bundles")) else "",
                }
        return details
    except Exception:
        return {}

def _load_strip_vendor_list() -> List[str]:
    """Load vendor names from StripSKUList.xlsx (for email tab)."""
    strip_path = Path(__file__).resolve().parent / STRIP_SKU_LIST_FILE
    if not strip_path.exists():
        return []
    try:
        df = pd.read_excel(strip_path, sheet_name="for email", header=0)
        if df.shape[0] == 0:
            return []
        vendor_col = _get_first_column(df, ["Vendor Name", "Vendor", "VendorName"])
        if not vendor_col:
            return []
        vendors = (
            df[vendor_col]
            .astype(str)
            .str.strip()
            .replace("nan", "")
            .replace("None", "")
        )
        vendors = sorted({v for v in vendors if v})
        return vendors
    except Exception:
        return []

def _load_veronica_payload() -> Dict:
    """Load flooring data filtered to only items in StripSKUList.xlsx"""
    original_data = _load_webapp_payload()
    strip_skus = _load_strip_sku_list()
    if not strip_skus:
        return original_data
    # Make a deep copy to avoid mutating the original data
    data = copy.deepcopy(original_data)
    # Filter items to only those in the strip SKU list
    items = data.get("items", [])
    filtered_items = [
        item for item in items
        if str(item.get("item_number", "")).strip().upper() in strip_skus
    ]
    data["items"] = filtered_items
    # Filter Inventory_Metrics
    metrics = data.get("Inventory_Metrics", [])
    filtered_metrics = [
        row for row in metrics
        if str(row.get("item_number", "") or row.get("sku", "")).strip().upper() in strip_skus
    ]
    data["Inventory_Metrics"] = filtered_metrics
    # Filter Monthly_Projections
    monthly = data.get("Monthly_Projections", [])
    filtered_monthly = [
        row for row in monthly
        if str(row.get("SKU", "")).strip().upper() in strip_skus
    ]
    data["Monthly_Projections"] = filtered_monthly
    # Filter arrivals
    arrivals = data.get("arrivals", [])
    filtered_arrivals = [
        row for row in arrivals
        if str(row.get("ITEM_NUMBER", "")).strip().upper() in strip_skus
    ]
    data["arrivals"] = filtered_arrivals
    return data

def _is_streamlit_runtime() -> bool:
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx() is not None
    except Exception:
        return False

def _build_forecast_chart(series: Dict) -> go.Figure:
    hist_x = series.get("hist_x") or series.get("x") or ["2025-11-23", "2025-11-30", "2025-12-07", "2025-12-14", "2025-12-21"]
    hist_vals = series.get("hist_y") or []
    fc_x = series.get("fc_x") or series.get("x") or ["2025-11-23", "2025-11-30", "2025-12-07", "2025-12-14", "2025-12-21"]
    fc_vals = series.get("fc_y") or series.get("y") or [72.4, 74.8, 74.9, 74.9, 74.1]

    hist_x = pd.to_datetime(hist_x, errors="coerce") if hist_x else hist_x
    fc_x = pd.to_datetime(fc_x, errors="coerce") if fc_x else fc_x

    fig = go.Figure()
    if hist_vals:
        fig.add_trace(
            go.Scatter(
                x=hist_x,
                y=hist_vals,
                mode="lines+markers",
                line=dict(color="#d46a1f", width=3),
                marker=dict(size=6, color="#d46a1f"),
                name="Historical",
            )
        )
    fig.add_trace(
        go.Scatter(
            x=fc_x,
            y=fc_vals,
            mode="lines+markers",
            line=dict(color="#2f6fdd", width=3),
            marker=dict(size=6, color="#2f6fdd"),
            name="Forecast",
        )
    )
    fig.update_layout(
        margin=dict(l=40, r=60, t=40, b=80),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="rgba(255,255,255,0.88)", family="Manrope, sans-serif"),
        showlegend=bool(hist_vals),
        legend=dict(font=dict(color="#ffffff"), orientation="h", x=0.5, y=-0.22, bgcolor="rgba(0,0,0,0)", xanchor="center", yanchor="top"),
        height=360,
    )
    fig.update_xaxes(
        showgrid=True,
        gridcolor="rgba(255,255,255,0.08)",
        ticks="",
        showline=False,
        tickfont=dict(size=11, color="rgba(255,255,255,0.65)"),
        tickformat="%b '%y",
    )
    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(255,255,255,0.08)",
        zeroline=False,
        tickfont=dict(size=11, color="rgba(255,255,255,0.65)"),
    )
    return fig

def _build_forecast_chart_multi(series_list: List[Dict]) -> go.Figure:
    fig = go.Figure()
    palette = ["#2f6fdd", "#6ccf7c", "#f7b84b", "#d96fd9", "#ff8b2c", "#66d0f5"]
    if not series_list:
        series_list = [{"label": "Demand", "series": {"x": ["Nov 23", "Nov 30"], "y": [72.4, 74.8]}}]

    for idx, item in enumerate(series_list):
        series = item.get("series", {})
        label = item.get("label", f"Item {idx+1}")
        color = palette[idx % len(palette)]

        # Plot historical data if available
        hist_x = series.get("hist_x") or []
        hist_y = series.get("hist_y") or []
        if hist_x and hist_y:
            fig.add_trace(
                go.Scatter(
                    x=hist_x,
                    y=hist_y,
                    mode="lines+markers",
                    line=dict(color=color, width=3),
                    marker=dict(size=6),
                    name=f"{label} (Historical)",
                    legendgroup=label,
                )
            )

        # Plot forecast data
        fc_x = series.get("fc_x") or series.get("x") or []
        fc_y = series.get("fc_y") or series.get("y") or []
        if fc_x and fc_y:
            fig.add_trace(
                go.Scatter(
                    x=fc_x,
                    y=fc_y,
                    mode="lines+markers",
                    line=dict(color=color, width=3, dash="dash"),
                    marker=dict(size=6),
                    name=f"{label} (Forecast)",
                    legendgroup=label,
                )
            )

    fig.update_layout(
        margin=dict(l=40, r=24, t=40, b=36),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="rgba(255,255,255,0.88)", family="Manrope, sans-serif"),
        showlegend=True,
        height=360,
        legend=dict(orientation="h", x=0.5, y=-0.22, xanchor="center", yanchor="top", font=dict(color="#ffffff")),
    )
    fig.update_xaxes(
        showgrid=True,
        gridcolor="rgba(255,255,255,0.08)",
        ticks="",
        showline=False,
        tickfont=dict(size=11, color="rgba(255,255,255,0.65)"),
    )
    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(255,255,255,0.08)",
        zeroline=False,
        tickfont=dict(size=11, color="rgba(255,255,255,0.65)"),
    )
    return fig

def _apply_legend_padding(fig: go.Figure, series_count: int) -> go.Figure:
    if series_count <= 6:
        legend_rows = 1
    else:
        legend_rows = int(math.ceil(series_count / 6))
    extra_bottom = max(0, (legend_rows - 1) * 22)
    base_height = 360
    base_bottom = 36
    fig.update_layout(
        height=base_height + extra_bottom,
        margin=dict(b=base_bottom + extra_bottom),
        legend=dict(y=-(0.22 + (legend_rows - 1) * 0.05)),
    )
    return fig

def _build_queue_df_from_inventory(metrics: List[Dict]) -> pd.DataFrame:
    if not metrics:
        return pd.DataFrame()
    df = pd.DataFrame(metrics)
    if "vendor_name" not in df.columns and "collection" in df.columns:
        df["vendor_name"] = df["collection"]
    rename_map = {
        "sku": "Item",
        "description": "Description",
        "vendor_name": "Vendor",
        "vendor_number": "Vend #",
        "available_sf": "Available",
        "on_po_sf": "On PO",
        "backorder_sf": "Backorder",
        "inventory_position": "Inv Position",
        "lead_time_days": "Lead Time (days)",
        "lead_time_mean_demand": "LT Demand",
        "safety_stock": "Safety Stock",
    }
    cols = [c for c in rename_map.keys() if c in df.columns]
    df = df[cols].rename(columns=rename_map)
    return df

def _render_queue_table_html(df: pd.DataFrame) -> str:
    if df is None or df.empty:
        return """
        <div class="queue-card">
          <div class="queue-header">INVENTORY AT A GLANCE</div>
          <div class="queue-empty">No items for selected vendor.</div>
        </div>
        """
    col_weights = {
        "Item": "0.9fr",
        "Description": "2.6fr",
        "Vendor": "1.3fr",
        "Vend #": "0.6fr",
        "Available": "1.0fr",
        "On PO": "0.7fr",
        "Backorder": "0.7fr",
        "Inv Position": "0.8fr",
        "Lead Time (days)": "0.9fr",
        "LT Demand": "0.9fr",
        "Safety Stock": "0.9fr",
    }
    headers = "".join(f"<div>{col}</div>" for col in df.columns)
    rows = []
    for _, row in df.iterrows():
        cells = []
        for col in df.columns:
            value = row[col]
            if isinstance(value, (int, float)) and col not in ("Item", "Description", "Vendor", "Vend #"):
                value = _format_number(float(value), 2)
            cells.append(f"<div class='text-clip'>{value}</div>")
        rows.append(f"<div class='queue-row'>{''.join(cells)}</div>")
    col_count = len(df.columns)
    col_template = " ".join(col_weights.get(col, "1fr") for col in df.columns)
    return f"""
    <div class="queue-card">
      <div class="queue-header">INVENTORY AT A GLANCE</div>
      <div class="queue-table" style="--col-count:{col_count}; --col-template:{col_template};">
        <div class="queue-row queue-head">{headers}</div>
        {''.join(rows)}
      </div>
    </div>
    """

def _render_reorder_now_table_html(df: pd.DataFrame, month_label: str) -> str:
    if df is None or df.empty:
        return f"""
        <div class="queue-card">
          <div class="queue-header">REORDER NOW: {month_label}</div>
          <div class="queue-empty">No reorder quantities for this month.</div>
        </div>
        """
    col_weights = {
        "Item Number": "1.1fr",
        "Collection": "1.1fr",
        "Description": "2.2fr",
        "Reorder Quantity (SF)": "1.2fr",
        "Reorder Quantity (Pallets)": "1.2fr",
        "Reorder Quantity (% of Container)": "1.4fr",
    }
    headers = "".join(f"<div>{col}</div>" for col in df.columns)
    rows = []
    for _, row in df.iterrows():
        cells = []
        for col in df.columns:
            value = row[col]
            if isinstance(value, (int, float)) and col in (
                "Reorder Quantity (SF)",
                "Reorder Quantity (Pallets)",
                "Reorder Quantity",
                "Quantity Needed",
            ):
                # Round up to next whole number, handle NaN
                if pd.isna(value):
                    value = ""
                else:
                    value = _format_number(math.ceil(float(value)), 0)
            elif isinstance(value, (int, float)) and col == "Reorder Quantity (% of Container)":
                if pd.isna(value):
                    value = ""
                else:
                    value = f"{float(value) * 100:.1f}%"
            elif value is None or (isinstance(value, float) and pd.isna(value)):
                value = ""
            cells.append(f"<div class='text-clip'>{value}</div>")
        rows.append(f"<div class='queue-row'>{''.join(cells)}</div>")
    col_count = len(df.columns)
    col_template = " ".join(col_weights.get(col, "1fr") for col in df.columns)
    return f"""
    <div class="queue-card">
      <div class="queue-header">REORDER NOW: {month_label}</div>
      <div class="queue-table" style="--col-count:{col_count}; --col-template:{col_template};">
        <div class="queue-row queue-head">{headers}</div>
        {''.join(rows)}
      </div>
    </div>
    """

def _render_reorder_table_html(df: pd.DataFrame, title: str) -> str:
    if df is None or df.empty:
        return f"""
        <div class="queue-card">
          <div class="queue-header">{title}</div>
          <div class="queue-empty">No reordering data available.</div>
        </div>
        """
    headers = "".join(f"<div>{col}</div>" for col in df.columns)
    rows = []
    for _, row in df.iterrows():
        cells = []
        for col in df.columns:
            value = row[col]
            if isinstance(value, (int, float)):
                if col == "Reorder Quantity" and abs(float(value)) < 1e-9:
                    value = ""
                else:
                    value = _format_number(float(value), 2)
            elif value is None:
                value = ""
            cells.append(f"<div class='text-clip'>{value}</div>")
        rows.append(f"<div class='queue-row'>{''.join(cells)}</div>")
    col_count = len(df.columns)
    label, details = title.split(":", 1) if ":" in title else ("REORDER SCHEDULE", title)
    parts = details.strip().split(" ", 1)
    sku = parts[0] if parts else ""
    desc = parts[1] if len(parts) > 1 else ""
    return f"""
    <div class="queue-card" style="max-width: 880px;">
      <div class="queue-header">
          <div class="reorder-header" style="--col-count:{col_count};">
          <div>{label}:</div>
          <div class="reorder-sku">{sku}</div>
          <div class="reorder-desc">{desc}</div>
        </div>
      </div>
      <div class="queue-table" style="--col-count:{col_count};">
        <div class="queue-row queue-head">{headers}</div>
        {''.join(rows)}
      </div>
    </div>
    """

def _render_arrivals_table_html(df: pd.DataFrame) -> str:
    if df is None or df.empty:
        return """
        <div class="queue-card">
          <div class="queue-header">ARRIVALS</div>
          <div class="queue-empty">No arrivals found for selection.</div>
        </div>
        """
    headers = "".join(f"<div>{col}</div>" for col in df.columns)
    rows = []
    for _, row in df.iterrows():
        cells = []
        for col in df.columns:
            value = row[col]
            if value is None:
                value = ""
            if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                value = ""
            if isinstance(value, str) and value.strip().lower() in ("nan", "none"):
                value = ""
            if isinstance(value, (int, float)) and col == "Quantity":
                value = _format_number(float(value), 2)
            cells.append(f"<div class='text-clip'>{value}</div>")
        rows.append(f"<div class='queue-row'>{''.join(cells)}</div>")
    col_weights = {
        "Item#": "0.9fr",
        "Description": "1.6fr",
        "PO#": "0.9fr",
        "Quantity": "0.8fr",
        "Container#": "2.6fr",
        "Due to Port": "1fr",
        "Due in Inventory": "1fr",
    }
    col_count = len(df.columns)
    col_template = " ".join(col_weights.get(col, "1fr") for col in df.columns)
    return f"""
    <div class="queue-card">
      <div class="queue-header">ARRIVALS</div>
      <div class="queue-table" style="--col-count:{col_count}; --col-template:{col_template};">
        <div class="queue-row queue-head">{headers}</div>
        {''.join(rows)}
      </div>
    </div>
    """

def _prepare_arrivals_df(
    arrivals_df: pd.DataFrame,
    selected_vendor: str,
    selected_collection: str,
    sku: Optional[str] = None,
) -> pd.DataFrame:
    if arrivals_df is None or arrivals_df.empty:
        return pd.DataFrame()
    df = arrivals_df.copy()
    col_po = _get_first_column(df, ["PO_NUMBER", "po_number", "PO#", "PLPO#"])
    col_item = _get_first_column(df, ["ITEM_NUMBER", "item_number", "PLITEM", "SKU"])
    col_desc = _get_first_column(df, ["DESCRIPTION", "description", "PLDESC"])
    col_qty = _get_first_column(df, ["QUANTITY_SF", "quantity_sf", "PLBLUO"])
    col_vendor = _get_first_column(df, ["VENDOR_NAME", "vendor_name", "Vendor"])
    col_collection = _get_first_column(df, ["COLLECTION", "collection"])
    col_due_port = _get_first_column(df, ["DUE_PORT", "due_port", "PLPDAT"])
    col_due_inv = _get_first_column(df, ["DUE_INV", "due_inv", "PLDDAT"])
    col_entry = _get_first_column(df, ["ENTRY_DATE", "entry_date", "PHDOI"])
    col_lt = _get_first_column(df, ["lead_time_days", "LEAD_TIME_DAYS"])
    col_container = _get_first_column(df, ["CONTAINER", "container"])
    if col_container is None:
        c1 = _get_first_column(df, ["PTCMT1", "ptcmt1"])
        c2 = _get_first_column(df, ["PTCMT2", "ptcmt2"])
        if c1 or c2:
            c1_vals = df[c1].fillna("").astype(str).str.strip() if c1 else ""
            c2_vals = df[c2].fillna("").astype(str).str.strip() if c2 else ""
            df["CONTAINER"] = np.where(c1_vals != "", c1_vals, c2_vals)
            col_container = "CONTAINER"

    if col_po is None:
        return pd.DataFrame()

    if col_item:
        df[col_item] = df[col_item].fillna("").astype(str).str.strip()
    if col_po:
        df[col_po] = df[col_po].fillna("").astype(str).str.strip()
    if col_item and col_po:
        keep = ~df[col_item].str.lower().isin(["", "nan", "none"])
        keep &= ~df[col_po].str.lower().isin(["", "nan", "none"])
        df = df[keep]

    if selected_vendor != "All Vendors" and col_vendor:
        df = df[df[col_vendor].astype(str).str.strip() == selected_vendor]
    if selected_collection != "All Collections" and col_collection:
        df = df[df[col_collection].astype(str).str.strip() == selected_collection]
    if sku and col_item:
        df = df[df[col_item].astype(str).str.strip() == str(sku)]

    dedupe_cols = [c for c in [col_po, col_item, col_due_port, col_due_inv, col_container] if c]
    if dedupe_cols:
        df = df.drop_duplicates(subset=dedupe_cols)

    container_series = df[col_container].fillna("").astype(str).str.strip() if col_container else ""
    container_series = container_series.replace("nan", "")
    due_port_series = df[col_due_port] if col_due_port else ""
    due_inv_series = df[col_due_inv] if col_due_inv else ""
    entry_series = df[col_entry] if col_entry else ""
    lt_series = df[col_lt] if col_lt else None

    ship_calc = []
    due_port_calc = []
    due_inv_calc = []
    for idx in range(len(df)):
        raw_due_inv = due_inv_series.iloc[idx] if col_due_inv else None
        raw_due_port = due_port_series.iloc[idx] if col_due_port else None
        raw_entry = entry_series.iloc[idx] if col_entry else None
        raw_ship = None
        lt_days = lt_series.iloc[idx] if lt_series is not None else None

        ship_dt = _normalize_arrival_date(raw_ship)
        due_port_dt = _normalize_arrival_date(raw_due_port)
        due_inv_dt = _normalize_arrival_date(raw_due_inv)
        entry_dt = _normalize_arrival_date(raw_entry)
        lt_days_val = float(lt_days) if lt_days is not None else None

        if pd.isna(ship_dt) and pd.isna(due_port_dt) and pd.isna(due_inv_dt):
            due_port_calc.append(pd.NaT)
            due_inv_calc.append(pd.NaT)
        elif not pd.isna(ship_dt) and pd.isna(due_port_dt) and pd.isna(due_inv_dt) and lt_days_val is not None:
            due_port_calc.append(_add_days_safe(ship_dt, max(lt_days_val - 21, 0)))
            due_inv_calc.append(_add_days_safe(ship_dt, lt_days_val))
        else:
            due_port_calc.append(raw_due_port)
            if pd.isna(due_inv_dt):
                if not pd.isna(due_port_dt):
                    due_inv_calc.append(_add_days_safe(due_port_dt, 21))
                elif not pd.isna(entry_dt) and lt_days_val is not None:
                    due_inv_calc.append(_add_days_safe(entry_dt, max(lt_days_val - 21, 0)))
                else:
                    due_inv_calc.append(pd.NaT)
            else:
                due_inv_calc.append(raw_due_inv)
        ship_calc.append(raw_ship)

    ship_calc = pd.Series(ship_calc, index=df.index)
    due_port_calc = pd.Series(due_port_calc, index=df.index)
    due_inv_calc = pd.Series(due_inv_calc, index=df.index)

    # Enforce chronological order: ship <= due port <= due inventory when dates exist.
    ship_norm = ship_calc.apply(_normalize_arrival_date)
    port_norm = due_port_calc.apply(_normalize_arrival_date)
    inv_norm = due_inv_calc.apply(_normalize_arrival_date)
    for i in range(len(df)):
        s = ship_norm.iloc[i]
        p = port_norm.iloc[i]
        inv = inv_norm.iloc[i]
        if pd.isna(s) and pd.isna(p) and pd.isna(inv):
            continue
        if pd.isna(s):
            s = p if not pd.isna(p) else inv
        if pd.isna(p):
            p = inv if not pd.isna(inv) else s
        if pd.isna(inv):
            inv = p if not pd.isna(p) else s
        if not pd.isna(s) and not pd.isna(p) and s > p:
            s = p
        if not pd.isna(p) and not pd.isna(inv) and p > inv:
            p = inv
        if not pd.isna(s) and not pd.isna(inv) and s > inv:
            s = inv
        ship_norm.iloc[i] = s
        port_norm.iloc[i] = p
        inv_norm.iloc[i] = inv
    item_series = df[col_item].fillna("").astype(str).str.strip() if col_item else ""
    desc_series = df[col_desc].fillna("").astype(str).str.strip() if col_desc else ""
    po_series = df[col_po].fillna("").astype(str).str.strip()
    out = pd.DataFrame(
        {
            "Item#": item_series.replace("nan", ""),
            "Description": desc_series.replace("nan", ""),
            "PO#": po_series.replace("nan", ""),
            "Quantity": df[col_qty] if col_qty else np.nan,
            "Container#": container_series,
            "Due to Port": port_norm.apply(_format_arrival_value),
            "Due in Inventory": inv_norm.apply(_format_arrival_value),
        }
    )
    for col in ["Item#", "Description", "PO#", "Container#"]:
        out[col] = out[col].fillna("").astype(str).str.strip()
        out[col] = out[col].replace({"nan": "", "None": ""})
    keep = ~out["Item#"].str.lower().isin(["", "nan", "none"])
    keep &= ~out["PO#"].str.lower().isin(["", "nan", "none"])
    out = out[keep].reset_index(drop=True)
    if col_due_inv:
        out["_sort"] = inv_norm
    out = out.sort_values("_sort", na_position="last").drop(columns=["_sort"])
    return out.reset_index(drop=True)

def _render_metric_card_html(title: str, value: str, subtitle: Optional[str] = None) -> str:
    subtitle_html = f"<div class='metric-sub'>{subtitle}</div>" if subtitle else ""
    return f"""
    <div class="metric-card">
      <div class="metric-title">{title}</div>
      <div class="metric-value">{value}</div>
      {subtitle_html}
    </div>
    """

def _render_segmentation_table_html(segmentation: Dict) -> str:
    rows = [
        ("maturity", segmentation.get("maturity", "")),
        ("demand_pattern", segmentation.get("demand_pattern", "")),
        ("ADI", _format_number(float(segmentation.get("adi", 0.0)), 6)),
        ("CV2", _format_number(float(segmentation.get("cv2", 0.0)), 6)),
        ("first_sales_date", segmentation.get("first_sales_date") or "--"),
        ("history_days", str(segmentation.get("history_days", 0))),
    ]
    body = "".join(
        f"<div class='seg-row'><div>{label}</div><div>{value}</div></div>" for label, value in rows
    )
    return f"""
    <div class="seg-card">
      <div class="seg-title">SEGMENTATION</div>
      {body}
    </div>
    """

def _aggregate_items(items: List[Dict]) -> Dict:
    if not items:
        return {}
    total_available = sum(float(item.get("available", 0.0) or 0.0) for item in items)
    total_on_po = sum(float(item.get("on_po", 0.0) or 0.0) for item in items)
    total_backorder = sum(float(item.get("backorder", 0.0) or 0.0) for item in items)
    total_inv = sum(float(item.get("inventory_position", 0.0) or 0.0) for item in items)
    total_lt_demand = sum(float(item.get("lt_demand", 0.0) or 0.0) for item in items)
    avg_daily = sum(float(item.get("daily_avg_demand", 0.0) or 0.0) for item in items)
    days_cover = total_inv / avg_daily if avg_daily > 0 else 0.0
    lead_time_days = int(round(np.mean([item.get("lead_time_days", 0) or 0 for item in items])))
    return {
        "available": total_available,
        "on_po": total_on_po,
        "backorder": total_backorder,
        "inventory_position": total_inv,
        "lt_demand": total_lt_demand,
        "daily_avg_demand": avg_daily,
        "days_of_cover": days_cover,
        "lead_time_days": lead_time_days,
        "vendor_number": items[0].get("vendor_number", "--"),
        "vendor_name": items[0].get("vendor_name", "--"),
    }

def _render_demand_graph_html(fig: go.Figure) -> str:
    fig_html = fig.to_html(include_plotlyjs=False, full_html=False, config={"displayModeBar": False})
    return f"""
    <html>
      <head>
        <script src="https://cdn.plot.ly/plotly-2.30.0.min.js"></script>
        <style>
          :root {{
            --panel-strong: rgba(30, 52, 16, 0.55);
            --panel: rgba(40, 63, 22, 0.55);
            --border: rgba(255,255,255,0.10);
            --text: #ffffff;
          }}
          body {{
            margin: 0;
            background: transparent;
            color: var(--text);
            font-family: "Manrope", sans-serif;
          }}
          .demand-card {{
            background: transparent;
            border-radius: 16px;
            overflow: hidden;
            border: 1px solid rgba(255,255,255,0.08);
          }}
          .demand-header {{
            background: var(--panel-strong);
            border-bottom: 1px solid rgba(255,255,255,0.08);
            padding: 10px 16px;
            font-size: 0.9rem;
            letter-spacing: 0.12em;
            font-weight: 700;
            color: var(--text);
            text-transform: uppercase;
          }}
          .demand-body {{
            background: var(--panel);
            padding: 12px 16px 8px 16px;
          }}
          .demand-body .plotly-graph-div {{
            margin: 0 !important;
          }}
        </style>
      </head>
      <body>
        <div class="demand-card">
          <div class="demand-header">DEMAND GRAPH</div>
          <div class="demand-body">
            {fig_html}
          </div>
        </div>
      </body>
    </html>
    """

def _series_from_monthly_rows(rows: List[Dict], sku: str) -> Dict:
    if not rows:
        return {}
    df = pd.DataFrame(rows)
    if df.empty or "SKU" not in df.columns or "Month" not in df.columns:
        return {}
    df = df[df["SKU"].astype(str) == str(sku)]
    if df.empty:
        return {}
    df["Month"] = pd.to_datetime(df["Month"], errors="coerce")
    df = df[df["Month"].notna()]
    if df.empty:
        return {}
    row_type_col = "Row_Type" if "Row_Type" in df.columns else None
    hist = df[df[row_type_col].astype(str).str.upper().eq("HIST")] if row_type_col else df[df["Historical Demand"].notna()]
    fc = df[df[row_type_col].astype(str).str.upper().isin(["FCST", "CATCHUP"])] if row_type_col else df[df["Forecast"].notna()]
    hist = hist.sort_values("Month")
    fc = fc.sort_values("Month")
    series: Dict[str, List[float]] = {}
    if not hist.empty and "Historical Demand" in hist.columns:
        series["hist_x"] = [d.strftime("%Y-%m-%d") for d in hist["Month"]]
        series["hist_y"] = [float(v) for v in hist["Historical Demand"]]
    if not fc.empty and "Forecast" in fc.columns:
        fc_series = fc.groupby("Month")["Forecast"].sum().sort_index()
        fc_series = _filter_forecast_months(fc_series)
        series["fc_x"] = [d.strftime("%Y-%m-%d") for d in fc_series.index]
        series["fc_y"] = [float(v) for v in fc_series.values]
        series["x"] = series["fc_x"]
        series["y"] = series["fc_y"]
    return series

def _series_from_monthly_rows_vendor(rows: List[Dict], vendor: str, key: str = "vendor_name") -> Dict:
    if not rows:
        return {}
    df = pd.DataFrame(rows)
    if "Month" not in df.columns:
        return {}
    # If key is None, aggregate all rows (no filtering)
    if key is not None:
        vendor_col = key if key in df.columns else ("vendor_name" if "vendor_name" in df.columns else "collection")
        if df.empty or vendor_col not in df.columns:
            return {}
        df = df[df[vendor_col].astype(str) == str(vendor)]
        if df.empty:
            return {}
    df["Month"] = pd.to_datetime(df["Month"], errors="coerce")
    df = df[df["Month"].notna()]
    if df.empty:
        return {}
    row_type_col = "Row_Type" if "Row_Type" in df.columns else None
    hist = df[df[row_type_col].astype(str).str.upper().eq("HIST")] if row_type_col else df[df["Historical Demand"].notna()]
    fc = df[df[row_type_col].astype(str).str.upper().isin(["FCST", "CATCHUP"])] if row_type_col else df[df["Forecast"].notna()]
    hist = hist.sort_values("Month")
    fc = fc.sort_values("Month")
    series: Dict[str, List[float]] = {}
    if not hist.empty and "Historical Demand" in hist.columns:
        hist_series = hist.groupby("Month")["Historical Demand"].sum().sort_index()
        series["hist_x"] = [d.strftime("%Y-%m-%d") for d in hist_series.index]
        series["hist_y"] = [float(v) for v in hist_series.values]
    if not fc.empty and "Forecast" in fc.columns:
        fc_series = fc.groupby("Month")["Forecast"].sum().sort_index()
        fc_series = _filter_forecast_months(fc_series)
        series["fc_x"] = [d.strftime("%Y-%m-%d") for d in fc_series.index]
        series["fc_y"] = [float(v) for v in fc_series.values]
        series["x"] = series["fc_x"]
        series["y"] = series["fc_y"]
    return series

def _build_spending_chart(spending: Dict) -> go.Figure:
    rng = np.random.default_rng(7)
    x_vals = spending.get("x") or [str(d) for d in range(1, 31)]
    this_month = spending.get("this") or []
    last_month = spending.get("last") or []

    if not this_month or not last_month:
        base = np.linspace(500, 16000, len(x_vals))
        noise = rng.normal(0, 260, len(x_vals)).cumsum()
        last_month = np.clip(base + noise, 0, None)
        this_month = np.clip(last_month * 0.68 + rng.normal(0, 220, len(x_vals)), 0, None)

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=x_vals,
            y=this_month,
            mode="lines",
            name="This month",
            line=dict(color="#f29a3a", width=4),
            fill="tozeroy",
            fillcolor="rgba(242,154,58,0.25)",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=x_vals,
            y=last_month,
            mode="lines",
            name="Last month",
            line=dict(color="#dfe8c6", width=4),
        )
    )
    title_text = spending.get("title", "Spending")
    subtitle = spending.get("subtitle", "This month vs. last month")
    fig.update_layout(
        title={
            "text": f"{title_text}<br><span style='font-size:0.8rem;color:rgba(241,244,228,0.75)'>{subtitle}</span>",
            "x": 0.02,
            "y": 0.94,
        },
        margin=dict(l=40, r=20, t=60, b=30),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="rgba(241,244,228,0.9)", family="Palatino Linotype, Book Antiqua, Palatino, serif"),
        legend=dict(orientation="h", yanchor="bottom", y=-0.22, x=0.2, font=dict(size=12)),
        showlegend=True,
        height=320,
    )
    show_ticks = bool(x_vals) and isinstance(x_vals[0], str)
    fig.update_xaxes(showgrid=False, ticks="", showticklabels=show_ticks)
    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(255,255,255,0.18)",
        zeroline=False,
        ticks="",
        tickprefix="$",
        tickformat=",.0f",
    )
    return fig

def _build_cashflow_chart(cashflow: Dict) -> go.Figure:
    months = cashflow.get("x") or ["Apr '25", "May '25", "Jun '25", "Jul '25"]
    inflow = np.array(cashflow.get("inflow") or [12000, 15000, 18000, 9000])
    outflow = np.array(cashflow.get("outflow") or [-17000, -16000, -14000, -9000])
    net = cashflow.get("net")
    if not net:
        net = (inflow + outflow).tolist()

    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=months,
            y=inflow,
            name="Inflow",
            marker_color="#3f7e57",
            opacity=0.9,
        )
    )
    fig.add_trace(
        go.Bar(
            x=months,
            y=outflow,
            name="Outflow",
            marker_color="#b94f45",
            opacity=0.9,
        )
    )
    fig.add_trace(
        go.Scatter(
            x=months,
            y=net,
            name="Net",
            mode="lines+markers",
            line=dict(color="#e6f0b1", width=3),
            marker=dict(size=6),
        )
    )
    fig.update_layout(
        title={"text": cashflow.get("title", "Cash flow trends"), "x": 0.04, "y": 0.95},
        margin=dict(l=40, r=20, t=60, b=30),
        barmode="relative",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="rgba(241,244,228,0.9)", family="Palatino Linotype, Book Antiqua, Palatino, serif"),
        showlegend=False,
        height=320,
    )
    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(255,255,255,0.18)",
        zeroline=False,
        tickprefix="$",
        tickformat=",.0f",
    )
    fig.update_xaxes(showgrid=False, ticks="", showline=False)
    return fig

def _build_transactions_html(transactions: List[Dict]) -> str:
    items = "\n".join(
        "<div class='list-row'>"
        f"<div class='list-left'><span class='list-icon'>{item.get('code','--')}</span>{item.get('name','--')}</div>"
        f"<div class='list-right'>{_format_currency(float(item.get('amount', 0.0)), 2)}</div>"
        "</div>"
        for item in transactions
    )
    return f"""
    <div class="card">
      <div class="card-title">Transactions</div>
      <div class="list">{items}</div>
    </div>
    """

def _build_categories_html(categories: List[Dict]) -> str:
    bars = []
    for item in categories:
        label = item.get("label", "--")
        value = _format_currency(float(item.get("value", 0.0)), 2)
        pct = float(item.get("pct", 0.0))
        color = item.get("color", "#5f8a4a")
        bars.append(
            f"""
            <div class="bar-row">
              <div class="bar-label">{label}</div>
              <div class="bar-track">
                <div class="bar-fill" style="width:{pct*100:.0f}%;background:{color};"></div>
              </div>
              <div class="bar-value">{value}</div>
            </div>
            """
        )
    return f"""
    <div class="card">
      <div class="card-title">Top expense categories</div>
      {''.join(bars)}
    </div>
    """

def _build_grocery_card_html(title: str, budget: float, actual: float, remaining: float, progress: float) -> str:
    progress_pct = max(0.0, min(1.0, progress))
    return f"""
    <div class="card">
      <div class="card-header-row">
        <div class="card-title">{title}</div>
        <div class="ring" style="--value:{progress_pct};"></div>
      </div>
      <div class="metric-grid">
        <div class="metric-block">
          <div class="metric-label">Budget</div>
          <div class="metric-value">{_format_currency(budget, 0)}</div>
        </div>
        <div class="metric-block">
          <div class="metric-label">Actual</div>
          <div class="metric-value">{_format_currency(actual, 0)}</div>
        </div>
        <div class="metric-block">
          <div class="metric-label">Remaining</div>
          <div class="metric-value positive">{_format_currency(remaining, 0)}</div>
        </div>
      </div>
    </div>
    """

def render_webapp(data: Optional[Dict] = None) -> None:
    st.set_page_config(page_title="OMP Purchasing Dashboard", layout="wide")
    if data is None:
        data = _load_webapp_payload()

    if "active_view" not in st.session_state:
        st.session_state.active_view = "Ilsy"

    st.markdown('<div class="toggle-wrap toggle-wrap--spaced">', unsafe_allow_html=True)
    switch_cols = st.columns([1, 1, 1], gap="medium")
    with switch_cols[0]:
        if st.button("Ilsy", use_container_width=True):
            st.session_state.active_view = "Ilsy"
    with switch_cols[1]:
        if st.button("Veronica", use_container_width=True):
            st.session_state.active_view = "Veronica"
    with switch_cols[2]:
        if st.button("Carlos", use_container_width=True):
            st.session_state.active_view = "Carlos"
    st.markdown("</div>", unsafe_allow_html=True)

    is_sundries = st.session_state.active_view == "Carlos"
    is_veronica = st.session_state.active_view == "Veronica"
    if is_sundries:
        data = _load_sundries_payload()
    elif is_veronica:
        data = _load_veronica_payload()
    else:
        data = _load_webapp_payload()

    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;600;700;800&display=swap');

        :root {
          --bg1: #6f8c34;
          --bg2: #5e7a2f;
          --bg3: #4f6927;
          --panel: rgba(40, 63, 22, 0.55);
          --panel-strong: rgba(30, 52, 16, 0.55);
          --panel-border: rgba(255,255,255,0.10);
          --text: #ffffff;
          --muted: #ffffff;
          --accent: #ff8a2a;
          --blue: #2f6fdd;
          --shadow: 0 12px 30px rgba(0,0,0,0.18);
          --shadow-soft: 0 6px 16px rgba(0,0,0,0.12);
          --section-header-size: 0.9rem;
        }

        html, body, [class*="css"]  {
          font-family: "Manrope", sans-serif;
          color: var(--text);
        }

        .stApp {
          background: linear-gradient(180deg, var(--bg1) 0%, var(--bg2) 60%, var(--bg3) 100%);
        }

        .block-container {
          padding-top: 1.4rem;
          padding-bottom: 2rem;
        }

        section[data-testid="stSidebar"] {
          background: linear-gradient(180deg, #ffffff 0 56px, rgba(30, 52, 16, 0.25) 56px);
          border-right: 1px solid rgba(255,255,255,0.08);
        }

        section[data-testid="stSidebar"] .stMarkdown {
          color: var(--text);
        }

        section[data-testid="stSidebar"] h3 {
          font-size: var(--section-header-size) !important;
        }

        section[data-testid="stSidebar"] label {
          color: #ffffff !important;
        }

        section[data-testid="stSidebar"] .stCheckbox label span {
          color: #ffffff !important;
        }

        section[data-testid="stSidebar"] .stCheckbox label p {
          color: #ffffff !important;
        }

        section[data-testid="stSidebar"] .stCheckbox div[role="checkbox"] + div span {
          color: #ffffff !important;
        }

        button[data-testid="collapsedControl"] {
          opacity: 1 !important;
          visibility: visible !important;
          display: inline-flex !important;
        }

        .hero-card {
          background: var(--panel);
          border: 1px solid var(--panel-border);
          border-radius: 18px;
          padding: 18px 20px;
          box-shadow: var(--shadow);
          margin-bottom: 18px;
        }

        .hero-top {
          display: flex;
          justify-content: space-between;
          align-items: flex-start;
          margin-bottom: 10px;
        }

        .pill {
          display: inline-flex;
          padding: 6px 12px;
          border-radius: 999px;
          border: 1px solid var(--panel-border);
          background: rgba(255,255,255,0.06);
          font-size: 0.75rem;
          letter-spacing: 0.08em;
          text-transform: uppercase;
          color: #ffffff;
        }

        .export-btn {
          display: inline-flex;
          align-items: center;
          justify-content: center;
          padding: 10px 14px;
          border-radius: 14px;
          border: 1px solid rgba(0,0,0,0.15);
          background: var(--accent);
          color: rgba(20,20,20,0.95);
          font-weight: 750;
          font-size: 0.8rem;
          text-decoration: none;
        }

        .hero-title {
          font-size: 1.7rem;
          font-weight: 800;
          margin: 8px 0;
          letter-spacing: 0.02em;
          color: #ffffff;
        }

        .hero-sub {
          color: #ffffff;
          font-size: 0.9rem;
          margin-top: 6px;
        }

        .hero-meta {
          text-align: right;
          color: #ffffff;
          font-size: 0.85rem;
          line-height: 1.35;
          opacity: 0.92;
          max-width: 360px;
        }

        .build-stamp {
          font-size: 0.72rem;
          letter-spacing: 0.06em;
          color: #ffffff;
          opacity: 0.8;
          margin-top: 6px;
        }

        .toggle-wrap {
          display: flex;
          gap: 12px;
          margin-bottom: 14px;
        }

        .toggle-wrap--spaced {
          margin-top: 32px;
        }

        .stButton button {
          background: rgba(0,0,0,0.25) !important;
          color: #ffffff !important;
          border: 1px solid rgba(255,255,255,0.25) !important;
          border-radius: 12px !important;
          font-weight: 700 !important;
        }

        .stButton button:focus {
          box-shadow: none !important;
        }

        .queue-card {
          background: var(--panel-strong);
          border-radius: 16px;
          padding: 0;
          margin-bottom: 18px;
          border: 1px solid rgba(255,255,255,0.08);
          box-shadow: var(--shadow);
        }

        .demand-body .plotly-html {
          margin: 0;
        }

        .queue-header {
          padding: 10px 16px;
          font-size: var(--section-header-size);
          letter-spacing: 0.12em;
          font-weight: 700;
          color: #ffffff;
          background: rgba(0,0,0,0.2);
          border-radius: 16px 16px 0 0;
        }

        .reorder-header {
          display: grid;
          grid-template-columns: repeat(var(--col-count, 5), minmax(0, 1fr));
          column-gap: 18px;
          align-items: center;
        }

        .reorder-desc {
          white-space: normal;
        }

        .queue-table {
          padding: 10px 12px 14px 12px;
        }

        .queue-row {
          display: grid;
          grid-template-columns: var(--col-template, repeat(var(--col-count, 5), minmax(0, 1fr)));
          column-gap: 18px;
          row-gap: 6px;
          font-size: 0.82rem;
          padding: 6px 8px;
          border-radius: 10px;
          color: rgba(255,255,255,0.98);
        }

        .queue-head {
          font-weight: 700;
          font-size: 0.74rem;
          letter-spacing: 0.03em;
          color: #ffffff;
          border-bottom: 1px solid rgba(255,255,255,0.08);
          margin-bottom: 6px;
          padding-bottom: 8px;
          text-transform: none;
        }

        .queue-row:not(.queue-head) {
          background: rgba(255,255,255,0.04);
        }

        .queue-empty {
          padding: 12px 16px 16px 16px;
          color: #ffffff;
          font-size: 0.85rem;
        }

        .optimizer-separator {
          height: 4px;
          background: #000000;
          border-radius: 999px;
          margin: 8px 4px 10px 4px;
        }

        div[data-testid="stVerticalBlock"]:has(.optimizer-anchor) .stMarkdown {
          margin-bottom: 4px;
        }

        div[data-testid="stVerticalBlock"]:has(.optimizer-anchor) input[aria-label="Description"] {
          font-size: 0.72rem;
        }

        div[data-testid="stVerticalBlock"]:has(.optimizer-anchor) div[data-testid="stCheckbox"] label {
          color: #ffffff;
          font-size: 0.92rem;
          font-weight: 600;
        }

        div[data-testid="stVerticalBlock"]:has(.optimizer-anchor) div[data-testid="stTextInput"],
        div[data-testid="stVerticalBlock"]:has(.optimizer-anchor) div[data-testid="stNumberInput"],
        div[data-testid="stVerticalBlock"]:has(.optimizer-anchor) div[data-testid="stSelectbox"] {
          margin-bottom: 6px;
        }

        .text-clip {
          white-space: nowrap;
          overflow: hidden;
          text-overflow: ellipsis;
        }

        .metric-stack {
          background: var(--panel);
          border-radius: 18px;
          padding: 14px 16px 12px 16px;
          border: 1px solid var(--panel-border);
          box-shadow: var(--shadow-soft);
        }

        .metric-stack .metric-title {
          margin-bottom: 6px;
        }

        .metric-row {
          display: flex;
          justify-content: space-between;
          align-items: baseline;
          padding: 6px 0;
          border-bottom: 1px solid rgba(255,255,255,0.08);
          font-size: 0.9rem;
        }

        .metric-row:last-child {
          border-bottom: none;
        }

        .metric-row span {
          color: #ffffff;
          font-size: 0.82rem;
        }

        .metric-card {
          background: var(--panel);
          border-radius: 18px;
          padding: 14px 16px 12px 16px;
          border: 1px solid var(--panel-border);
          box-shadow: var(--shadow-soft);
          margin-bottom: 14px;
        }

        .metric-title {
          font-size: 0.72rem;
          letter-spacing: 0.12em;
          color: #ffffff;
          text-transform: uppercase;
        }

        .metric-value {
          font-size: 1.55rem;
          font-weight: 750;
          margin-top: 6px;
          color: rgba(255,255,255,0.96);
        }

        .metric-sub {
          font-size: 0.78rem;
          color: #ffffff;
          margin-top: 6px;
        }

        .seg-card {
          background: #f4f5ee;
          color: #2a2f16;
          border-radius: 14px;
          padding: 12px 14px;
          box-shadow: var(--shadow);
          border: 1px solid rgba(0,0,0,0.08);
        }

        .seg-title {
          font-size: 0.75rem;
          font-weight: 800;
          color: #5b682a;
          letter-spacing: 0.1em;
          margin-bottom: 10px;
        }

        .seg-row {
          display: grid;
          grid-template-columns: 1fr 1fr;
          padding: 6px 0;
          border-bottom: 1px solid rgba(0,0,0,0.08);
          font-size: 0.8rem;
        }

        .seg-row:last-child {
          border-bottom: none;
        }

        div[data-testid="stPlotlyChart"] {
          background: transparent;
          border: 0;
          border-radius: 0;
          padding: 12px 0 4px 0;
          box-shadow: none;
          overflow: hidden;
        }

        .demand-header + div[data-testid="stPlotlyChart"] {
          padding-right: 0;
        }

        </style>
        """,
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.markdown("### Item Filters")

    items = data.get("items", []) or _demo_webapp_payload().get("items", [])
    if not items:
        items = _demo_webapp_payload().get("items", [])

    vendor_names = sorted({str(item.get("vendor_name", "")).strip() for item in items if item.get("vendor_name")})
    if not vendor_names:
        vendor_names = ["--"]
    vendor_choices = ["All Vendors"] + vendor_names
    selected_vendor = st.sidebar.selectbox("Vendor name", vendor_choices, index=0)

    if selected_vendor == "All Vendors":
        vendor_items = items
    else:
        vendor_items = [item for item in items if str(item.get("vendor_name", "")).strip() == selected_vendor]

    if is_sundries:
        selected_collection = "All Collections"
        collection_items = vendor_items
    else:
        collection_names = sorted({str(item.get("collection", "")).strip() for item in vendor_items if item.get("collection")})
        if not collection_names:
            collection_names = ["--"]
        collection_choices = ["All Collections"] + collection_names
        selected_collection = st.sidebar.selectbox("Collection", collection_choices, index=0)

        if selected_collection == "All Collections":
            collection_items = vendor_items
        else:
            collection_items = [
                item for item in vendor_items if str(item.get("collection", "")).strip() == selected_collection
            ]

    if not collection_items:
        collection_items = vendor_items
    item_number_options = ["All items"] + [
        str(item.get("item_number", "")).strip()
        for item in collection_items
        if str(item.get("item_number", "")).strip()
    ]
    selected_item_number = st.sidebar.selectbox("Item number", item_number_options, index=0)

    if selected_item_number != "All items":
        collection_items = [
            item for item in collection_items if str(item.get("item_number", "")).strip() == selected_item_number
        ]

    item_options = ["All items"] + [
        (str(item.get("description", "")).strip() or str(item.get("item_number", "")).strip())
        for item in collection_items
    ]
    selected_desc = st.sidebar.selectbox("Item description", item_options, index=0)

    if selected_desc != "All items":
        selected_item = next(
            (item for item in collection_items if str(item.get("description", "")) == selected_desc),
            collection_items[0],
        )
    elif selected_item_number != "All items":
        selected_item = collection_items[0] if collection_items else None
    else:
        selected_item = None
    vendor_items = collection_items
    aggregate_graphs = st.sidebar.checkbox(
        "Aggregate Graphs",
        value=False,
        help="Checked = aggregate demand by vendor. Unchecked = show one line per SKU.",
    )

    run_meta = data.get("run_meta", {})
    if selected_item:
        header_title = f"{selected_item.get('item_number','')} | {selected_item.get('description','')}"
        vendor_label = selected_item.get("vendor_name") or "--"
        vendor_number = selected_item.get("vendor_number") or "--"
        vendor_summary = selected_item
    else:
        header_title = f"{selected_vendor} | All items"
        vendor_label = selected_vendor or "--"
        vendor_summary = _aggregate_items(vendor_items)
        vendor_number = vendor_summary.get("vendor_number", "--")
    horizon_days = run_meta.get("forecast_horizon_days", FUTURE_FORECAST_DAYS)
    run_stamp = run_meta.get("run_timestamp_local", "--")

    if is_sundries:
        title_text = "OMP Sundries Purchasing Dashboard"
    elif is_veronica:
        title_text = "OMP Strip Flooring Purchasing Dashboard"
    else:
        title_text = "OMP Engineered Floor Purchasing Dashboard"
    st.markdown(
        f"""
        <div class="hero-card">
          <div class="hero-top">
            <div class="pill">{title_text}</div>
            <div class="hero-meta">
              <div>{vendor_label} (Vendor {vendor_number})</div>
              <div>Horizon: {horizon_days} days | Run: {run_stamp}</div>
              <div>UI build: {UI_BUILD}</div>
            </div>
          </div>
          <div class="hero-title">{header_title}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    metrics_rows = data.get("Inventory_Metrics", [])
    if selected_collection != "All Collections":
        metrics_rows = [
            row for row in metrics_rows if str(row.get("collection", "")).strip() == selected_collection
        ]
    if selected_item:
        item_id = str(selected_item.get("item_number", "")).strip()
        if item_id:
            metrics_rows = [
                row for row in metrics_rows if str(row.get("item_number", "")).strip() == item_id
            ]
    queue_df = _build_queue_df_from_inventory(metrics_rows)
    if not queue_df.empty and selected_vendor != "All Vendors":
        queue_df = queue_df[queue_df["Vendor"].astype(str).str.strip() == selected_vendor]
    if selected_item and not queue_df.empty and "Item" in queue_df.columns:
        queue_df = queue_df[queue_df["Item"].astype(str).str.strip() == str(selected_item.get("item_number", "")).strip()]
    if queue_df.empty:
        queue_df = _build_queue_df_from_inventory(_demo_webapp_payload().get("Inventory_Metrics", []))

    monthly_rows = data.get("Monthly_Projections", [])
    arrivals_rows = data.get("arrivals", [])
    if selected_collection != "All Collections":
        monthly_rows = [
            row for row in monthly_rows if str(row.get("collection", "")).strip() == selected_collection
        ]
    if selected_item:
        series = selected_item.get("forecast_series", {}) or {}
        if not series.get("hist_y"):
            series = {**series, **_series_from_monthly_rows(monthly_rows, selected_item.get("item_number", ""))}
        forecast_fig = _build_forecast_chart(series)
    else:
        if selected_vendor != "All Vendors":
            if aggregate_graphs:
                series_key = "collection" if selected_collection != "All Collections" else "vendor_name"
                series_value = selected_collection if series_key == "collection" else selected_vendor
                vendor_series = _series_from_monthly_rows_vendor(monthly_rows, series_value, key=series_key)
                if vendor_series.get("hist_y") or vendor_series.get("fc_y"):
                    forecast_fig = _build_forecast_chart(vendor_series)
                else:
                    # Fallback: aggregate from item forecast_series with historical data
                    series_list = []
                    for item in vendor_items:
                        sku = item.get("item_number", "")
                        item_series = _series_from_monthly_rows(monthly_rows, sku)
                        fallback_series = item.get("forecast_series", {}) or {}
                        if not item_series.get("hist_y") and fallback_series.get("hist_y"):
                            item_series["hist_x"] = fallback_series.get("hist_x")
                            item_series["hist_y"] = fallback_series.get("hist_y")
                        if not item_series.get("fc_y") and (fallback_series.get("fc_y") or fallback_series.get("y")):
                            item_series["fc_x"] = fallback_series.get("fc_x") or fallback_series.get("x")
                            item_series["fc_y"] = fallback_series.get("fc_y") or fallback_series.get("y")
                    series_list.append({
                        "label": item.get("item_number", item.get("description", "")),
                        "series": item_series,
                    })
                    forecast_fig = _build_forecast_chart_multi(series_list)
                    forecast_fig = _apply_legend_padding(forecast_fig, len(series_list))
            else:
                series_list = []
                for item in vendor_items:
                    sku = item.get("item_number", "")
                    series = _series_from_monthly_rows(monthly_rows, sku)
                    fallback_series = item.get("forecast_series", {}) or {}
                    if not series.get("hist_y") and fallback_series.get("hist_y"):
                        series["hist_x"] = fallback_series.get("hist_x")
                        series["hist_y"] = fallback_series.get("hist_y")
                    if not series.get("fc_y") and (fallback_series.get("fc_y") or fallback_series.get("y")):
                        series["fc_x"] = fallback_series.get("fc_x") or fallback_series.get("x")
                        series["fc_y"] = fallback_series.get("fc_y") or fallback_series.get("y")
                    series_list.append(
                        {
                            "label": item.get("item_number", item.get("description", "")),
                            "series": series,
                        }
                    )
                forecast_fig = _build_forecast_chart_multi(series_list)
                forecast_fig = _apply_legend_padding(forecast_fig, len(series_list))
        else:
            # All Vendors selected
            if aggregate_graphs:
                if selected_collection != "All Collections":
                    # Aggregate by collection when only collection is selected
                    collection_series = _series_from_monthly_rows_vendor(monthly_rows, selected_collection, key="collection")
                else:
                    # Aggregate all items (for Veronica tab with All Vendors / All Collections)
                    collection_series = _series_from_monthly_rows_vendor(monthly_rows, None, key=None)
                if collection_series.get("hist_y") or collection_series.get("fc_y"):
                    forecast_fig = _build_forecast_chart(collection_series)
                else:
                    # Fallback: aggregate from item forecast_series
                    series_list = []
                    for item in vendor_items:
                        sku = item.get("item_number", "")
                        item_series = _series_from_monthly_rows(monthly_rows, sku)
                        fallback_series = item.get("forecast_series", {}) or {}
                        if not item_series.get("hist_y") and fallback_series.get("hist_y"):
                            item_series["hist_x"] = fallback_series.get("hist_x")
                            item_series["hist_y"] = fallback_series.get("hist_y")
                        if not item_series.get("fc_y") and (fallback_series.get("fc_y") or fallback_series.get("y")):
                            item_series["fc_x"] = fallback_series.get("fc_x") or fallback_series.get("x")
                            item_series["fc_y"] = fallback_series.get("fc_y") or fallback_series.get("y")
                        series_list.append({
                            "label": item.get("item_number", item.get("description", "")),
                            "series": item_series,
                        })
                    forecast_fig = _build_forecast_chart_multi(series_list)
                    forecast_fig = _apply_legend_padding(forecast_fig, len(series_list))
            else:
                # Show each item with historical data
                series_list = []
                for item in vendor_items:
                    sku = item.get("item_number", "")
                    item_series = _series_from_monthly_rows(monthly_rows, sku)
                    fallback_series = item.get("forecast_series", {}) or {}
                    if not item_series.get("hist_y") and fallback_series.get("hist_y"):
                        item_series["hist_x"] = fallback_series.get("hist_x")
                        item_series["hist_y"] = fallback_series.get("hist_y")
                    if not item_series.get("fc_y") and (fallback_series.get("fc_y") or fallback_series.get("y")):
                        item_series["fc_x"] = fallback_series.get("fc_x") or fallback_series.get("x")
                        item_series["fc_y"] = fallback_series.get("fc_y") or fallback_series.get("y")
                    series_list.append({
                        "label": item.get("item_number", item.get("description", "")),
                        "series": item_series,
                    })
                forecast_fig = _build_forecast_chart_multi(series_list)
                forecast_fig = _apply_legend_padding(forecast_fig, len(series_list))

    fig_html = forecast_fig.to_html(include_plotlyjs="cdn", full_html=False)
    demand_html = _render_demand_graph_html(forecast_fig)
    fig_height = int(getattr(forecast_fig.layout, "height", 360) or 360)
    components.html(demand_html, height=fig_height + 120, scrolling=False)

    arrivals_df_all = pd.DataFrame(arrivals_rows)
    arrivals_all_out = _prepare_arrivals_df(arrivals_df_all, selected_vendor, selected_collection)
    if selected_item and not arrivals_all_out.empty and "Item#" in arrivals_all_out.columns:
        arrivals_all_out = arrivals_all_out[
            arrivals_all_out["Item#"].astype(str).str.strip()
            == str(selected_item.get("item_number", "")).strip()
        ]
    st.markdown(_render_arrivals_table_html(arrivals_all_out), unsafe_allow_html=True)
    st.markdown(_render_queue_table_html(queue_df), unsafe_allow_html=True)

    # Reorder Now section (current month, Order Quantity > 0)
    reorder_now_df = pd.DataFrame()
    optimizer_source_df = pd.DataFrame()
    if monthly_rows:
        df_monthly_all = pd.DataFrame(monthly_rows)
        if "Month" in df_monthly_all.columns and "Order Quantity" in df_monthly_all.columns:
            df_monthly_all["Month"] = pd.to_datetime(df_monthly_all["Month"], errors="coerce")
            current_month = pd.Timestamp.today().to_period("M").to_timestamp()
            df_current = df_monthly_all[df_monthly_all["Month"] == current_month]
            df_current = df_current[pd.to_numeric(df_current["Order Quantity"], errors="coerce") > 0]
            if selected_vendor != "All Vendors" and "vendor_name" in df_current.columns:
                df_current = df_current[df_current["vendor_name"].astype(str).str.strip() == selected_vendor]
            if selected_collection != "All Collections" and "collection" in df_current.columns:
                df_current = df_current[df_current["collection"].astype(str).str.strip() == selected_collection]
            if selected_item and "SKU" in df_current.columns:
                df_current = df_current[df_current["SKU"].astype(str) == str(selected_item.get("item_number", ""))]

            if not df_current.empty:
                optimizer_source_df = df_current.copy()
                reorder_sf = pd.to_numeric(df_current["Order Quantity"], errors="coerce")
                if is_sundries:
                    reorder_now_df = pd.DataFrame(
                        {
                            "Item Number": df_current.get("SKU", ""),
                            "Collection": df_current.get("collection", ""),
                            "Description": df_current.get("description", ""),
                            "Reorder Quantity": reorder_sf,
                        }
                    )
                    total_sf = reorder_now_df["Reorder Quantity"].sum(skipna=True)
                    total_row = {
                        "Item Number": "Total",
                        "Collection": "",
                        "Description": "",
                        "Reorder Quantity": total_sf,
                    }
                    reorder_now_df = pd.concat(
                        [reorder_now_df, pd.DataFrame([total_row])], ignore_index=True
                    )
                elif is_veronica:
                    # Load Strip SKU details for Veronica tab
                    strip_details = _load_strip_sku_details()
                    item_numbers = df_current.get("SKU", pd.Series()).astype(str).str.strip().str.upper()

                    # Get details for each item
                    thickness_list = [strip_details.get(sku, {}).get("Thickness", "") for sku in item_numbers]
                    width_list = [strip_details.get(sku, {}).get("Width", "") for sku in item_numbers]
                    species_list = [strip_details.get(sku, {}).get("Species", "") for sku in item_numbers]
                    grade_cut_list = [strip_details.get(sku, {}).get("Grade/Cut", "") for sku in item_numbers]
                    edge_list = [strip_details.get(sku, {}).get("Edge", "") for sku in item_numbers]
                    bundles_list = [strip_details.get(sku, {}).get("Bundles", "") for sku in item_numbers]

                    reorder_now_df = pd.DataFrame(
                        {
                            "Item Number": df_current.get("SKU", ""),
                            "Thickness": thickness_list,
                            "Width": width_list,
                            "Species": species_list,
                            "Grade/Cut": grade_cut_list,
                            "Edge": edge_list,
                            "Bundles": bundles_list,
                            "Quantity Needed": reorder_sf,
                            "Price": "",
                            "Freight": "",
                            "Additional Charges": "",
                        }
                    )
                    total_qty = reorder_now_df["Quantity Needed"].sum(skipna=True)
                    total_row = {
                        "Item Number": "Total",
                        "Thickness": "",
                        "Width": "",
                        "Species": "",
                        "Grade/Cut": "",
                        "Edge": "",
                        "Bundles": "",
                        "Quantity Needed": total_qty,
                        "Price": "",
                        "Freight": "",
                        "Additional Charges": "",
                    }
                    reorder_now_df = pd.concat(
                        [reorder_now_df, pd.DataFrame([total_row])], ignore_index=True
                    )
                else:
                    sf_per_pallet_col = df_current.get("sf_per_pallet")
                    pallets_per_container_col = df_current.get("pallets_per_container")

                    # Handle case where columns may not exist
                    if sf_per_pallet_col is not None:
                        sf_per_pallet = pd.to_numeric(sf_per_pallet_col, errors="coerce").replace(0, np.nan)
                    else:
                        sf_per_pallet = pd.Series([np.nan] * len(df_current), index=df_current.index)

                    if pallets_per_container_col is not None:
                        pallets_per_container = pd.to_numeric(pallets_per_container_col, errors="coerce").replace(0, np.nan)
                    else:
                        pallets_per_container = pd.Series([np.nan] * len(df_current), index=df_current.index)

                    reorder_pallets = reorder_sf / sf_per_pallet
                    reorder_pct = reorder_pallets / pallets_per_container

                    reorder_now_df = pd.DataFrame(
                        {
                            "Item Number": df_current.get("SKU", ""),
                            "Collection": df_current.get("collection", ""),
                            "Description": df_current.get("description", ""),
                            "Reorder Quantity (SF)": reorder_sf,
                            "Reorder Quantity (Pallets)": reorder_pallets,
                            "Reorder Quantity (% of Container)": reorder_pct,
                        }
                    )
                    total_sf = reorder_now_df["Reorder Quantity (SF)"].sum(skipna=True)
                    total_pallets = reorder_now_df["Reorder Quantity (Pallets)"].sum(skipna=True)
                    total_pct = reorder_now_df["Reorder Quantity (% of Container)"].sum(skipna=True)
                    total_row = {
                        "Item Number": "Total",
                        "Collection": "",
                        "Description": "",
                        "Reorder Quantity (SF)": total_sf,
                        "Reorder Quantity (Pallets)": total_pallets,
                        "Reorder Quantity (% of Container)": total_pct,
                    }
                    reorder_now_df = pd.concat(
                        [reorder_now_df, pd.DataFrame([total_row])], ignore_index=True
                    )
    month_label = pd.Timestamp.today().strftime("%B %Y")
    st.markdown(_render_reorder_now_table_html(reorder_now_df, month_label), unsafe_allow_html=True)

    # Export Reorder Now to Excel button
    if not reorder_now_df.empty:
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        # Create Excel file in memory
        excel_buffer = io.BytesIO()
        # Remove the Total row for export
        export_df = reorder_now_df[reorder_now_df["Item Number"] != "Total"].copy()

        # Round up Quantity Needed to next decimal point (1 decimal place)
        if "Quantity Needed" in export_df.columns:
            export_df["Quantity Needed"] = export_df["Quantity Needed"].apply(
                lambda x: math.ceil(x * 10) / 10 if pd.notna(x) else x
            )

        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name='Reorder Now', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Reorder Now']

            # Define column widths
            col_widths = {
                "Item Number": 20,
                "Thickness": 10,
                "Width": 8,
                "Species": 15,
                "Grade/Cut": 15,
                "Edge": 12,
                "Bundles": 10,
                "Quantity Needed": 20,
                "Price": 8,
                "Freight": 8,
                "Additional Charges": 18,
            }

            # Define styles
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # Header: dark blue text, light blue fill (60% lighter)
            header_font = Font(bold=True, color="1F4E79")  # Dark blue text
            header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Light blue fill

            # Apply column widths
            for col_idx, col_name in enumerate(export_df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                width = col_widths.get(col_name, 12)
                worksheet.column_dimensions[col_letter].width = width

            # Apply header formatting (row 1)
            for col_idx in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # Apply borders to all data cells
            for row_idx in range(2, len(export_df) + 2):
                for col_idx in range(1, len(export_df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border

        excel_buffer.seek(0)
        file_date = pd.Timestamp.today().strftime("%Y-%m-%d")
        st.download_button(
            label="Export Reorder Now to Excel",
            data=excel_buffer,
            file_name=f"reorder_now_{file_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _optimizer_mark_include(key_suffix: str) -> None:
        include_key = f"optimizer_include_{key_suffix}"
        any_filled = False
        for idx in range(1, 10):
            vendor_val = st.session_state.get(f"optimizer_vendor_{key_suffix}_{idx}", "")
            price_val = st.session_state.get(f"optimizer_price_{key_suffix}_{idx}", "")
            freight_val = st.session_state.get(f"optimizer_freight_{key_suffix}_{idx}", "")
            fees_val = st.session_state.get(f"optimizer_fees_{key_suffix}_{idx}", "")
            if str(vendor_val).strip() or str(price_val).strip() or str(freight_val).strip() or str(fees_val).strip():
                any_filled = True
                break
        st.session_state[include_key] = any_filled

    vendor_options = _load_strip_vendor_list()
    if not vendor_options:
        vendor_options = [""]
    optimizer_rows = []
    strip_details = _load_strip_sku_details() if is_veronica else {}
    if not reorder_now_df.empty:
        source_rows = reorder_now_df[reorder_now_df["Item Number"] != "Total"].copy()
        for _, row in source_rows.iterrows():
            sku = str(row.get("Item Number", "")).strip()
            desc = str(row.get("Description", "")).strip()
            if not desc and is_veronica:
                desc = strip_details.get(sku, {}).get("Description", "")
            if not desc and is_veronica:
                desc_parts = []
                for col in ("Thickness", "Width", "Species", "Grade/Cut", "Edge", "Bundles"):
                    val = row.get(col)
                    if val is None:
                        continue
                    text = str(val).strip()
                    if text and text.lower() not in ("nan", "none"):
                        desc_parts.append(text)
                if desc_parts:
                    desc = " / ".join(desc_parts)
            qty_value = None
            for col in (
                "Quantity Needed",
                "Reorder Quantity (SF)",
                "Reorder Quantity",
                "Reorder Quantity (Pallets)",
            ):
                if col in source_rows.columns:
                    qty_value = row.get(col)
                    break
            qty_text = ""
            if isinstance(qty_value, (int, float)) and not pd.isna(qty_value):
                qty_text = _format_number(float(qty_value), 2)
            elif isinstance(qty_value, str):
                qty_text = qty_value
            optimizer_rows.append(
                {
                    "sku": sku,
                    "description": desc,
                    "quantity": qty_text,
                }
            )

    optimizer_container = st.container()
    with optimizer_container:
        st.markdown('<div class="optimizer-anchor"></div>', unsafe_allow_html=True)
        st.markdown('<div class="queue-header">OPTIMIZER</div>', unsafe_allow_html=True)

        if optimizer_rows:
            total_rows = len(optimizer_rows)
            for row_idx, row in enumerate(optimizer_rows, start=1):
                key_suffix = "".join(ch if ch.isalnum() else "_" for ch in row["sku"]) or str(row_idx)
                cols = st.columns([1.4] + [1.0] * 9, gap="small")
                with cols[0]:
                    st.text_input(
                        "SKU",
                        key=f"optimizer_sku_{key_suffix}",
                        value=row["sku"],
                        label_visibility="collapsed",
                    )
                    st.text_input(
                        "Description",
                        key=f"optimizer_description_{key_suffix}",
                        value=row["description"],
                        label_visibility="collapsed",
                    )
                    st.text_input(
                        "Quantity Needed",
                        key=f"optimizer_qty_needed_{key_suffix}",
                        value=row["quantity"],
                        label_visibility="collapsed",
                    )
                    include_key = f"optimizer_include_{key_suffix}"
                    if include_key not in st.session_state:
                        st.session_state[include_key] = False
                    st.checkbox(
                        "Include",
                        key=include_key,
                    )
                for idx in range(1, 10):
                    with cols[idx]:
                        st.selectbox(
                            f"Vendor {idx}",
                            vendor_options,
                            key=f"optimizer_vendor_{key_suffix}_{idx}",
                            index=None if vendor_options != [""] else 0,
                            placeholder=f"Vendor {idx}",
                            label_visibility="collapsed",
                            on_change=_optimizer_mark_include,
                            args=(key_suffix,),
                        )
                        st.text_input(
                            f"Price {idx}",
                            key=f"optimizer_price_{key_suffix}_{idx}",
                            placeholder=f"Price {idx}",
                            label_visibility="collapsed",
                            on_change=_optimizer_mark_include,
                            args=(key_suffix,),
                        )
                        st.text_input(
                            f"Freight {idx}",
                            key=f"optimizer_freight_{key_suffix}_{idx}",
                            placeholder=f"Freight {idx}",
                            label_visibility="collapsed",
                            on_change=_optimizer_mark_include,
                            args=(key_suffix,),
                        )
                        st.text_input(
                            f"Fees {idx}",
                            key=f"optimizer_fees_{key_suffix}_{idx}",
                            placeholder=f"Fees {idx}",
                            label_visibility="collapsed",
                            on_change=_optimizer_mark_include,
                            args=(key_suffix,),
                        )
                if row_idx < total_rows:
                    st.markdown('<div class="optimizer-separator"></div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="queue-empty">No reorder quantities for this month.</div>', unsafe_allow_html=True)

    # Monthly detail table under Purchasing Queue
    detail_items = [selected_item] if selected_item else vendor_items
    if detail_items:
        df_monthly_all = pd.DataFrame(monthly_rows)
        for item in detail_items:
            if not item:
                continue
            sku_label = item.get("item_number", "")
            desc_label = item.get("description", "")
            df_monthly = df_monthly_all.copy()
            if not df_monthly.empty and "SKU" in df_monthly.columns:
                df_monthly = df_monthly[df_monthly["SKU"].astype(str) == str(sku_label)]
            # Filter out HIST rows - only show CATCHUP and FCST rows with forecast data
            if not df_monthly.empty and "Row_Type" in df_monthly.columns:
                df_monthly = df_monthly[df_monthly["Row_Type"].astype(str).str.upper().isin(["FCST", "CATCHUP"])]
            if not df_monthly.empty and "Month" in df_monthly.columns:
                df_monthly["Month"] = pd.to_datetime(df_monthly["Month"], errors="coerce")
                df_monthly = df_monthly.sort_values("Month")
                df_monthly["Month"] = df_monthly["Month"].dt.strftime("%m/%d/%Y")
            col_map = {
                "Month": "Month",
                "Beginning Inventory": "Beginning Inventory",
                "Forecast": "Forecast",
                "Order Quantity": "Reorder Quantity",
                "Ending Inventory": "Ending Inventory",
            }
            existing = [col for col in col_map if col in df_monthly.columns]
            if existing:
                out = df_monthly[existing].rename(columns=col_map)
                if "Beginning Inventory" in out.columns:
                    out["Beginning Inventory"] = pd.to_numeric(out["Beginning Inventory"], errors="coerce").fillna(0.0)
                if "Ending Inventory" in out.columns:
                    out["Ending Inventory"] = pd.to_numeric(out["Ending Inventory"], errors="coerce").fillna(0.0)
                st.markdown(
                    _render_reorder_table_html(out, f"Reorder Schedule:   {sku_label} {desc_label}"),
                    unsafe_allow_html=True,
                )

    # Removed metric/segmentation sections below Purchasing Queue per request.
if __name__ == "__main__":
    if _is_streamlit_runtime():
        render_webapp()
        raise SystemExit
    print("="*70)
    print("INVENTORY PLANNING - DEMAND-AWARE FORECASTING")
    print("="*70)
    print(f"Mode: {'Single SKU - ' + FOCUS_SKU if FOCUS_SKU else 'All Active SKUs'}")
    print(f"Date Filter: Only SKUs with sales since {CUTOFF_DATE}")
    print(f"Data Aggregation: WEEKLY (reduces intermittency, improves accuracy)")
    print(f"Historical Data: Each SKU uses data from its first sale month forward")
    print(f"\nService Level: {SERVICE_LEVEL*100:.0f}%")
    print(f"\nService Level: {SERVICE_LEVEL*100:.0f}%")
    print(f"Safety Stock Cap: Maximum 3 months of average demand")
    print(f"\nForecast Method Selection:")
    print(f"  ALL 9+ models tested per SKU (Neural Network, XGBoost, Random Forest,")
    print(f"  Holt Damped, SARIMA, TSB, SBA, Croston, Hurdle, + Global models)")
    print(f"  Best method selected based on wMAPE on validation data")
    print(f"\nDemand Classification (ADI/CV²): SMOOTH, INTERMITTENT, ERRATIC, LUMPY")
    print(f"\nABC Uplift Scalars: A={SS_UPLIFT_SCALAR['A']:.1f}, B={SS_UPLIFT_SCALAR['B']:.1f}, C={SS_UPLIFT_SCALAR['C']:.1f}")
    print(f"Global Uplift Budget: {GLOBAL_UPLIFT_BUDGET_PCT*100:.0f}% of trailing 12m volume")
    
    if ENABLE_GLOBAL_INV_CAP:
        print(f"\n{'='*70}")
        print("GLOBAL INVENTORY CAP ENABLED")
        print(f"{'='*70}")
        print(f"  Max Total Ending Inventory: {GLOBAL_INV_CAP_SF:,.0f} SF")
        print(f"  ABC MOI Caps: A={ABC_MOI_CAPS['A']:.1f}, B={ABC_MOI_CAPS['B']:.1f}, C={ABC_MOI_CAPS['C']:.1f} months")
        print(f"  Strategy: Min (lead time demand) + Flex (scaled to fit capacity)")
    
    print("="*70)
    run_inventory_planning()
